# Keep only these imports (external modules)
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from .models import Media
from .serializers import MediaSerializer
from .permissions import IsSystemAdminOrReadOnly
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.permissions import IsAdminUser, AllowAny, SAFE_METHODS, BasePermission
from django.http import JsonResponse
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.db.models import Avg, Count, Sum
from rest_framework.decorators import action
from .serializers import LogoutSerializer
from drf_spectacular.utils import extend_schema
from django.shortcuts import get_object_or_404
from django.contrib.auth import authenticate, get_user_model
from django.contrib.auth import logout

from rest_framework import status, permissions, viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import NotFound
from drf_spectacular.utils import extend_schema, OpenApiParameter
from .serializers import MFAPasswordVerifySerializer
from .utils.mfa_token import create_mfa_settings_token, verify_mfa_settings_token
from .serializers import MFAPasswordVerifySerializer, MFAToggleSerializer

# from .models import OnboardingState
from .models import CrisisHotline
from .serializers import CrisisHotlineSerializer

# from .serializers import OnboardingStateSerializer
from rest_framework.permissions import (
    IsAuthenticated,
    BasePermission,
    IsAuthenticatedOrReadOnly,
    AllowAny,
)

import django_filters
from drf_spectacular.utils import extend_schema_view, extend_schema
from django.db.models import Avg

from .models import Feedback
from .serializers import FeedbackSerializer
import logging
import csv
import os

logger = logging.getLogger(__name__)
from django.db.models import Avg
from rest_framework import viewsets, permissions, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from drf_spectacular.types import OpenApiTypes
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, CharFilter
from drf_spectacular.utils import extend_schema, extend_schema_view

from .models import Feedback
from .serializers import FeedbackSerializer
import django_filters
from .models import Feedback
from .models import DynamicQuestion
from .models import AssessmentQuestion
from .serializers import AssessmentQuestionSerializer
from .serializers import DynamicQuestionSerializer
import random
from django.db import transaction
from rest_framework import viewsets, permissions, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter

from rest_framework.generics import RetrieveUpdateAPIView, UpdateAPIView
from rest_framework.exceptions import ValidationError
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import *
from .serializers import (
    EmployeeInvitationCreateSerializer,
    EmployeeInvitationResponseSerializer,
)
from django.core.mail import send_mail, EmailMultiAlternatives
from .utils.gmail_http_api import send_gmail_api_email
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
from datetime import datetime, timedelta
from django.utils.timezone import now
import secrets
from rest_framework import filters
import string
import pyotp, qrcode, io, base64
from django.utils.crypto import get_random_string
from django.core.cache import cache
import calendar
from .models import Organization, ChatSession, ChatMessage, EmployeeProfile
from .serializers import (
    PasswordResetOTPVerificationSerializer,
    InvitationOTPVerificationSerializer,
    ChatSessionSerializer,
    ChatMessageSerializer,
)
from .serializers import OrganizationCreateSerializer
from django.template.loader import render_to_string
from django.contrib.auth import authenticate, login as django_login
from .serializers import AdminUserSerializer, OrganizationSerializer
from rest_framework_simplejwt.exceptions import TokenError
from .permissions import IsSystemAdmin
import hmac
import hashlib
from openpyxl import Workbook
import requests
from .models import FeatureUsage
from .serializers import FeatureUsageSerializer
from .utils.feature import FeatureUsageCalculator




from django.views.decorators.csrf import csrf_exempt  #  Make sure this is present
from django.http import HttpResponse
from django.conf import settings
from rest_framework.decorators import api_view, permission_classes
import logging
from django_filters.rest_framework import DjangoFilterBackend
from .serializers import EmployeeProfileSerializer
# from sana_ai.services.mental_health_ai import get_ai_service  # TODO: Add sana_ai app
from .Services.groq_service import GroqService

# AI Status Management imports
from .models import AIStatus
from .serializers import AIStatusSerializer


# Mood score mapping for trends (used to provide numeric scores to frontend)
MOOD_SCORES = {
    "Ecstatic": 5,
    "Happy": 4,
    "Excited": 4,
    "Content": 4,
    "Calm": 3,
    "Neutral": 3,
    "Tired": 3,
    "Anxious": 2,
    "Stressed": 2,
    "Sad": 2,
    "Frustrated": 2,
    "Angry": 1,
}


# Set up logging
logging.getLogger(__name__)

# Get User model
User = get_user_model()


# Helper function (Moved from a 'services' file into views.py)
def verify_flutterwave_transaction(tx_ref, expected_amount, currency):
    """Verifies a transaction using the Flutterwave API."""
    url = f"https://api.flutterwave.com/v3/transactions/{tx_ref}/verify"
    headers = {
        "Authorization": f"Bearer {settings.FLW_SEC_KEY}",
        "Content-Type": "application/json",
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()

        if data["status"] == "success" and data["data"]["status"] == "successful":
            # Security Check: Compare amount and currency to prevent tampering
            if (
                data["data"]["amount"] >= expected_amount
                and data["data"]["currency"] == currency
            ):
                return True, data["data"]  # Verification success
            else:
                return False, {"detail": "Amount or currency mismatch."}
        else:
            return False, {"detail": "Transaction failed or not found."}
    except requests.exceptions.RequestException as e:
        return False, {"detail": f"Flutterwave API verification error: {e}"}


def initiate_payment_fw(
    amount, email, subscription_id, currency="NGN"
):  # Renamed transaction_id to payment_ref
    url = "https://api.flutterwave.com/v3/payments"
    headers = {"Authorization": f"Bearer {settings.FLW_SEC_KEY}"}
    tx_ref = str(uuid.uuid4())

    data = {
        "tx_ref": tx_ref,
        "amount": str(amount),
        "currency": currency,
        "redirect_url": f"http://64.225.122.101:8000/api/v1/billing/verify_payment/?sub_ref={subscription_id}",
        # "redirect_url": f"{FRONTEND_SUCCESS_URL}?tx_ref={tx_ref}",
        "meta": {
            "subscription_id": subscription_id,  # Pass  internal reference
        },
        "customer": {"email": email, "name": "user_name"},
        "customizations": {
            "title": "Subscription Payment",
            "description": "Payment for subscription plan",
            "logo": "http://www.piedpiper.com/app/themes/joystick-v27/images/logo.png",
        },
    }


# Permission: company admin (is_staff)
class IsCompanyAdmin(BasePermission):
    """Allows access only to users with is_staff=True."""

    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated)


# # SIGNUP VIEW
@extend_schema(tags=["Authentication"])
class SignupView(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = SignupSerializer
    permission_classes = [permissions.AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        refresh = RefreshToken.for_user(user)

        return Response({
            "message": "Employee account created successfully",
            "user": {
                "id": user.id,
                "email": user.email,
                "role": user.role,
            },
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "onboarding_required": True,
            "redirect_url": "/onboarding/"
        }, status=status.HTTP_201_CREATED)

# VIEWS FOR CREATING AN ORGANIZATION
@extend_schema(
    tags=["Authentication"],
    request=OrganizationCreateSerializer,
    responses={201: OrganizationCreateSerializer},
)
class OrganizationSignupView(viewsets.ModelViewSet):
    queryset = Organization.objects.all()
    serializer_class = OrganizationCreateSerializer
    permission_classes = [permissions.AllowAny]


# view for organization details
class OrganizationDetailView(APIView):
    def get(self, request, org_id):
        try:
            org = Organization.objects.get(id=org_id)
        except Organization.DoesNotExist:
            return Response({"error": "Organization not found"}, status=404)

        serializer = OrganizationDetailSerializer(org)
        return Response(serializer.data)


# Verify reset password otp view
@extend_schema(
    tags=["Authentication"],
    request=PasswordResetOTPVerificationSerializer,
    responses={200: "OTP verified"},
    description="Verifies the OTP sent to the user's email for password reset",
)
class VerifyPasswordResetOTPView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = PasswordResetOTPVerificationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        otp = serializer.context["otp"]
        otp.delete()  # Remove OTP so it cannot be reused

        return Response(
            {"message": "OTP verified successfully. You can now reset your password."},
            status=status.HTTP_200_OK,
        )


# Verify invitation otp view
@extend_schema(
    tags=["Authentication"],
    request=InvitationOTPVerificationSerializer,
    responses={200: "Invitation OTP verified"},
    description="Verifies the invitation OTP before employee account creation",
)
class VerifyInvitationOTPView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = InvitationOTPVerificationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        invitation = serializer.context["invitation"]

        # Mark invitation as accepted
        invitation.accepted_at = timezone.now()
        invitation.save()

        # Save verified invitation email in session
        request.session["verified_invitation_email"] = invitation.email
        request.session["invitation_verified_at"] = timezone.now().isoformat()

        return Response(
            {
                "message": "Invitation OTP verified successfully. Proceed to create your account.",
                "email": invitation.email,
                "employer": invitation.employer.name,
            },
            status=status.HTTP_200_OK,
        )


# LOGIN VIEW - Build success payload
def _build_login_success_payload(user):
    refresh = RefreshToken.for_user(user)

    user_data = {
        "id": user.id,
        "username": user.email,  # Use email as username since USERNAME_FIELD = 'email'
        "email": user.email,
        "role": user.role,
        "date_joined": user.date_joined,
        "is_active": user.is_active,
        "avatar": user.avatar.url if hasattr(user, "avatar") and user.avatar else None,
        "onboarding_completed": getattr(user, "onboarding_completed", False),
    }

    # Redirect URL based on role
    if user.role == "system_admin":
        redirect_url = "/admin/dashboard/"
    elif user.role == "employer":
        redirect_url = "/organization/dashboard/"
    elif user.role == "employee":
        redirect_url = "/api/v1/mobile-login-success/"
    else:
        redirect_url = "/"

    return {
        "refresh": str(refresh),
        "access": str(refresh.access_token),
        "user": user_data,
        "redirect_url": redirect_url,
    }


# LOGIN VIEW - Main login endpoint
@extend_schema(
    request=LoginSerializer,
    responses={200: OpenApiTypes.OBJECT},
    tags=['Authentication'],
    description="Login using email and password. MFA integrated if enabled."
)
class LoginView(APIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = LoginSerializer
    queryset = None

    def post(self, request):
        serializer = self.serializer_class(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)

        user = serializer.validated_data["user"]

        # This is MFA Check
        if user.mfa_enabled:
            temp_token = get_random_string(32)
            cache.set(temp_token, user.id, timeout=300)  # valid 5 minutes
            
            # Always return MFA setup data for first-time login
            if user.mfa_secret:
                totp = pyotp.TOTP(user.mfa_secret)
                otpauth_uri = totp.provisioning_uri(name=user.email, issuer_name="Obeeoma System Admin")
                
                # Generate QR code
                qr = qrcode.make(otpauth_uri)
                buffer = io.BytesIO()
                qr.save(buffer, format="PNG")
                qr_b64 = base64.b64encode(buffer.getvalue()).decode()
                
                return Response({
                    "mfa_required": True, 
                    "temp_token": temp_token,
                    "mfa_setup_data": {
                        "qr_code_base64": qr_b64,
                        "secret": user.mfa_secret,
                        "otpauth_uri": otpauth_uri
                    }
                })
            
            return Response({"mfa_required": True, "temp_token": temp_token})

        # Onboarding required ONLY for employees
        if user.role == "employee" and not getattr(user, "onboarding_completed", False):
            return Response(
                {
                    "onboarding_required": True,
                    "temp_access_token": str(RefreshToken.for_user(user).access_token),
                    "message": "Onboarding required before using the system.",
                },
                status=200,
            )

        # Login normally
        django_login(request, user)
        return Response(_build_login_success_payload(user))


# Matching view for custom token obtain pair serializer
@extend_schema(
    tags=["Authentication"],
    request=CustomTokenObtainPairSerializer,
    responses={200: CustomTokenObtainPairSerializer},
)
class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


#  CompleteOnboardingView
class CompleteOnboardingView(APIView):
    """
    Completes first-time employee onboarding.
    User must be authenticated but NOT onboarded.
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        request=EmployeeOnboardingSerializer,
        responses={200: OpenApiTypes.OBJECT},
        tags=["Onboarding"],
        description="Complete first-time user onboarding.",
    )
    @transaction.atomic
    def post(self, request):
        user = request.user

        # this Prevents re-onboarding
        if user.onboarding_completed:
            return Response(
                {"detail": "Onboarding already completed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate onboarding payload
        serializer = EmployeeOnboardingSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Complete onboarding
        serializer.update(user, serializer.validated_data)

        # this Persists onboarding state (UX tracking)
        OnboardingState.objects.update_or_create(
            user=user, defaults={"completed": True, "first_action_done": True}
        )

        # this Ensures that the DB is up-to-date
        user.refresh_from_db()

        return Response(
            {
                "message": "Onboarding completed successfully.",
                "onboarding_completed": user.onboarding_completed,
                "first_time_access": False,
            },
            status=status.HTTP_200_OK,
        )


#  MarkOnboardingCompleteView
class MarkOnboardingCompleteView(APIView):
    """
    Marks user onboarding as completed without requiring additional data.
    
    PURPOSE:
    - Frontend saves data incrementally (avatar, assessments separately)
    - Original complete-onboarding endpoint requires password + all data at once
    - This endpoint only sets the onboarding_completed flag to True
    
    USE CASE:
    - Called after frontend has already saved all onboarding data separately
    - User is already authenticated (no password needed)
    - Prevents duplicate assessment creation
    - Solves the issue where onboarding_completed never gets set to True
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        responses={200: OpenApiTypes.OBJECT},
        tags=["Onboarding"],
        description="Mark onboarding as complete after incremental saves.",
    )
    @transaction.atomic
    def post(self, request):
        """
        Sets onboarding_completed = True for authenticated user.
        
        This endpoint is designed to be called AFTER the frontend has:
        1. Already saved the user's avatar through separate endpoint
        2. Already saved all assessments through separate endpoints
        3. User is already authenticated (no password required)
        
        The original complete-onboarding endpoint cannot be used because:
        - It requires password (user already logged in)
        - It would create duplicate assessments
        - It requires all data at once (frontend saves incrementally)
        """
        user = request.user

        # Check if onboarding is already completed
        if user.onboarding_completed:
            return Response(
                {
                    "message": "Onboarding already completed.",
                    "onboarding_completed": True,
                },
                status=status.HTTP_200_OK,
            )

        # Mark onboarding as completed
        user.onboarding_completed = True
        user.is_first_time = False
        
        # Save only the onboarding-related fields
        user.save(update_fields=[
            "onboarding_completed",
            "is_first_time"
        ])

        # Update onboarding state tracking (for UX analytics)
        OnboardingState.objects.update_or_create(
            user=user, 
            defaults={
                "completed": True, 
                "first_action_done": True
            }
        )

        # Ensure database is up-to-date
        user.refresh_from_db()

        return Response(
            {
                "message": "Onboarding marked as completed successfully.",
                "onboarding_completed": user.onboarding_completed,
                "is_first_time": user.is_first_time,
            },
            status=status.HTTP_200_OK,
        )
    
# Feature Usage ViewSet
class FeatureUsageViewSet(viewsets.ModelViewSet):
    """
    Handles listing, creating, and tracking feature usage
    """
    serializer_class = FeatureUsageSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Users only see their own usage
        return FeatureUsage.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        feature_name = serializer.validated_data['feature']
        # Increment existing or create new
        usage = FeatureUsageCalculator.track_feature(self.request.user, feature_name)
        serializer.instance = usage

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        GET /feature-usage/stats/
        Returns feature usage percentages for the logged-in user
        """
        data = FeatureUsageCalculator.calculate_percentage_for_user(request.user)
        return Response(data)


# LOGOUT VIEW
@extend_schema(
    tags=["Authentication"],
    request=LogoutSerializer,
    responses={
        200: {"description": "Logged out successfully"},
        400: {"description": "Invalid refresh token"},
        501: {"description": "Token blacklist not enabled"},
    },
)
class LogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        serializer = LogoutSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        refresh_token = serializer.validated_data["refresh"]
        try:
            token = RefreshToken(refresh_token)
            token.blacklist()  # requires token_blacklist app installed
        except AttributeError:
            # Blacklist app not installed
            return Response(
                {"detail": "Token blacklist not enabled on server."},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )
        except TokenError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {"message": "Logged out successfully"}, status=status.HTTP_200_OK
        )


# password reset request view
@extend_schema(tags=["Authentication"])
class PasswordResetView(viewsets.ViewSet):
    permission_classes = [permissions.AllowAny]
    serializer_class = PasswordResetSerializer

    def create(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data.get("email")

        user = User.objects.filter(email=email).first()
        if not user:
            return Response(
                {
                    "message": f"If an account exists for {email}, a reset code has been sent."
                },
                status=status.HTTP_200_OK,
            )

        try:
            code = "".join(secrets.choice(string.digits) for _ in range(6))
            token = secrets.token_urlsafe(32)
            expires_at = timezone.now() + timedelta(minutes=5)

            PasswordResetToken.objects.filter(user=user).delete()

            reset_token = PasswordResetToken.objects.create(
                user=user, token=token, code=code, expires_at=expires_at
            )
        except Exception as e:
            logger.error("Error generating password reset token: %s", str(e))
            return Response(
                {"error": "Something went wrong while generating reset token."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # Send email using Gmail API with HTML template
        subject = "Password Reset Verification Code - Obeeoma"
        
        # Render HTML template
        html_content = render_to_string('emails/reset.html', {
            'user_email': email,
            'user_name': user.username,
            'otp_code': code,
            'otp_expiry': '5 minutes'
        })
        
        # Plain text version for fallback
        message = f"""
Hello {user.username},

You requested a password reset for your Obeeoma account.

Your verification code is: {code}

This code will expire in 5 minutes.

If you did not request this password reset, please ignore this email.

Best regards,
Obeeoma Team
"""

        try:
            # Add detailed logging
            logger.info(f"Attempting to send password reset email to: {email}")
            logger.info(f"Subject: {subject}")
            logger.info(f"User: {user.username}")
            logger.info(f"Reset code: {code}")
            
            success = send_gmail_api_email(email, subject, message, html_body=html_content)
            
            if success:
                logger.info(f"Password reset email sent successfully to {email}")
            else:
                logger.error(f"Failed to send password reset email to {email}")
                
            if not success:
                raise Exception("Gmail API failed to send email")

            return Response(
                {
                    "message": f"If an account exists for {email}, a reset code has been sent.",
                    "token": token,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            logger.error("Error sending password reset email: %s", str(e))
            reset_token.delete()
            return Response(
                {"error": "Failed to send email. Please try again later."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# Confirm Password Reset View
@extend_schema(tags=["Authentication"])
class PasswordResetConfirmView(viewsets.ViewSet):
    permission_classes = [permissions.AllowAny]
    serializer_class = PasswordResetSerializer

    def create(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        code = serializer.validated_data["code"]
        new_password = serializer.validated_data["new_password"]

        # Get token from request headers or body
        token = request.data.get("token")
        if not token:
            return Response(
                {"error": "Token is required"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # This part checks if token and code are valid
            reset_token = PasswordResetToken.objects.get(
                token=token, code=code, is_used=False
            )

            # This part checks if token is expired
            if reset_token.is_expired():
                reset_token.delete()
                return Response(
                    {"error": "Verification code has expired"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # This helps in Updating user password
            user = reset_token.user
            user.set_password(new_password)
            user.save()

            # Mark token as used
            reset_token.mark_as_used()

            return Response(
                {"message": "Password reset successfully"}, status=status.HTTP_200_OK
            )

        except PasswordResetToken.DoesNotExist:
            return Response(
                {"error": "Invalid verification code"},
                status=status.HTTP_400_BAD_REQUEST,
            )


# View for changing or updating password
@extend_schema(
    tags=["Authentication"],
    request=PasswordChangeSerializer,
    responses={200: {"message": "Password updated successfully. Please log in again."}},
    description="Allows a logged-in user to change their password. Optionally accepts 'refresh' token to blacklist it.",
)
class PasswordChangeView(viewsets.ViewSet):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def create(self, request):
        serializer = PasswordChangeSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # Optional: blacklist refresh token
        refresh_token = request.data.get("refresh")
        if refresh_token:
            try:
                token = RefreshToken(refresh_token)
                token.blacklist()
            except Exception:
                pass

        return Response(
            {"message": "Password updated successfully. Please log in again."},
            status=status.HTTP_200_OK,
        )


# This is the Setup for MFA (when the superuser is already logged in)
@extend_schema(request=MFASetupSerializer, responses={200: MFASetupSerializer})
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mfa_setup(request):
    user = request.user
    if not (user.role == "system_admin" or user.is_superuser):
        return Response(
            {"error": "Only system administrators can enable MFA"}, status=403
        )

    raw_secret = user.generate_mfa_secret()
    totp = pyotp.TOTP(raw_secret)
    otpauth_uri = totp.provisioning_uri(name=user.username, issuer_name="ObeeomaApp")

    # This helps the system to Generate the QR image in base64
    qr = qrcode.make(otpauth_uri)
    buffer = io.BytesIO()
    qr.save(buffer, format="PNG")
    qr_b64 = base64.b64encode(buffer.getvalue()).decode()

    return Response(
        {
            "otpauth_uri": otpauth_uri,
            "qr_code_base64": qr_b64,
            "secret": raw_secret,
        }
    )


# This setup comfirms the MFA


@extend_schema(request=MFAConfirmSerializer, responses={200: MFAConfirmSerializer})
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mfa_confirm(request):
    user = request.user
    code = request.data.get("code")

    secret = user.get_mfa_secret()
    if not secret:
        return Response({"error": "MFA not initialized"}, status=400)

    totp = pyotp.TOTP(secret)
    if not totp.verify(code, valid_window=1):
        return Response({"error": "Invalid verification code"}, status=400)

    user.mfa_enabled = True
    user.save(update_fields=["mfa_enabled"])
    return Response({"detail": "MFA successfully enabled"})


# This logic is for MFA Verification during login
# So this logic will help in MFA verification by checking the code
@extend_schema(request=MFAVerifySerializer, responses={200: MFAVerifySerializer})
@api_view(["POST"])
@permission_classes([AllowAny])
def mfa_verify(request):
    serializer = MFAVerifySerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    temp_token = serializer.validated_data["temp_token"]
    code = serializer.validated_data["code"]

    user_id = cache.get(temp_token)
    if not user_id:
        return Response({"error": "Expired or invalid session"}, status=400)

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        cache.delete(temp_token)
        return Response(
            {"error": "Account associated with this session was not found"}, status=400
        )

    if not user.mfa_enabled:
        cache.delete(temp_token)
        return Response({"error": "MFA is not enabled for this account"}, status=400)

    secret = user.get_mfa_secret()
    if not secret:
        return Response({"error": "MFA is not configured for this account"}, status=400)

    totp = pyotp.TOTP(secret)
    if not totp.verify(code, valid_window=1):
        return Response({"error": "Invalid or expired code"}, status=401)

    # This logic says if Valid code ,finalize login
    cache.delete(temp_token)
    django_login(request, user)  # This sets the Django session cookie

    return Response(_build_login_success_payload(user))


# Verify Password for MFA Actions
@extend_schema(tags=["MFA"])
class VerifyMFAPasswordView(APIView):
    serializer_class = MFAPasswordVerifySerializer
    permission_classes = [IsAuthenticated]

    @extend_schema(
        description="Verify admin password before allowing MFA changes. Returns a temporary MFA settings token."
    )
    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        password = serializer.validated_data["password"]

        if not user.check_password(password):
            return Response({"error": "Incorrect password."}, status=400)

        token = create_mfa_settings_token(user.id)
        return Response(
            {"message": "Password verified successfully.", "mfa_settings_token": token}
        )


# MFA Toggle View
@extend_schema(tags=["MFA"])
class ToggleMFAView(APIView):
    serializer_class = MFAToggleSerializer
    permission_classes = [IsAuthenticated]

    @extend_schema(
        description="Enable or disable MFA securely using the temporary MFA settings token."
    )
    def patch(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        mfa_enabled = serializer.validated_data["mfa_enabled"]
        token = serializer.validated_data["mfa_settings_token"]

        user_id = verify_mfa_settings_token(token)
        if not user_id or user_id != request.user.id:
            return Response(
                {"error": "Invalid or expired MFA settings token."}, status=400
            )

        user = request.user
        user.mfa_enabled = mfa_enabled
        user.save()

        return Response(
            {
                "message": "MFA setting updated successfully.",
                "mfa_enabled": user.mfa_enabled,
            }
        )


# Resetpassword completeview
@extend_schema(tags=["Authentication"])
class ResetPasswordCompleteView(viewsets.ViewSet):
    permission_classes = [AllowAny]
    serializer_class = ResetPasswordCompleteSerializer

    def create(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = serializer.validated_data["user"]
        new_password = serializer.validated_data["new_password"]

        user.set_password(new_password)
        user.save()

        return Response(
            {"message": "Password has been reset successfully."},
            status=status.HTTP_200_OK,
        )


class EmployeeUserCreateSerializer(serializers.ModelSerializer):
    password = serializers.CharField(write_only=True, min_length=8)
    password_confirm = serializers.CharField(write_only=True)

    class Meta:
        model = User
        fields = ["email", "first_name", "last_name", "password", "password_confirm"]

    def validate(self, attrs):
        if attrs["password"] != attrs["password_confirm"]:
            raise serializers.ValidationError(
                {"password_confirm": "Passwords don't match"}
            )
        return attrs

    def create(self, validated_data):
        validated_data.pop("password_confirm")
        user = User.objects.create_user(
            username=validated_data["email"],  # Use email as username
            email=validated_data["email"],
            password=validated_data["password"],
            first_name=validated_data["first_name"],
            last_name=validated_data["last_name"],
            is_active=True,
        )
        return user


# VIEWS FOR HOTLINE
class ActiveHotlineView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        hotline = CrisisHotline.objects.filter(is_active=True).first()
        serializer = CrisisHotlineSerializer(hotline)
        return Response(serializer.data)


# --- Employee Invitation Views ---
@extend_schema(tags=["Employee Invitations"])
class InviteView(viewsets.ModelViewSet):
    """
    Employee Invitation Management (OTP Only)

    Allows employers to:
    - Send email invitations to new employees
    - View pending / accepted / expired invitations
    """

    serializer_class = EmployeeInvitationCreateSerializer
    permission_classes = [permissions.IsAuthenticated]

    # QUERYSET (LIST / FILTER)

    def get_queryset(self):
        employer = None
        user = self.request.user
        now = timezone.now()

        # Get employer from organization
        organization = Organization.objects.filter(contact_person=user).first()
        if organization:
            employer, _ = Employer.objects.get_or_create(
                name=organization.organizationName, defaults={"is_active": True}
            )

        # Get employer from employee profile
        if not employer:
            employee_profile = Employee.objects.filter(user=user).first()
            if employee_profile:
                employer = employee_profile.employer

        # Staff fallback
        if not employer and user.is_staff:
            employer = Employer.objects.first()

        if not employer:
            return EmployeeInvitation.objects.none()

        queryset = EmployeeInvitation.objects.filter(employer=employer).order_by(
            "-created_at"
        )

        status_param = self.request.query_params.get("status")

        if status_param == "pending":
            queryset = queryset.filter(
                accepted_at__isnull=True,
                rejected_at__isnull=True,
                otp_expires_at__gt=now,
            )

        elif status_param == "accepted":
            queryset = queryset.filter(accepted_at__isnull=False)

        elif status_param == "expired":
            queryset = queryset.filter(
                accepted_at__isnull=True,
                rejected_at__isnull=True,
                otp_expires_at__lte=now,
            )

        return queryset

    # CREATE INVITATION
    @extend_schema(
        request=EmployeeInvitationCreateSerializer,
        responses={201: EmployeeInvitationResponseSerializer},
        description="Send an invitation to a new employee with OTP",
    )
    def create(self, request, *args, **kwargs):
        user = request.user
        employer = None

        # Get employer from organization
        organization = Organization.objects.filter(contact_person=user).first()
        if organization:
            employer, _ = Employer.objects.get_or_create(
                name=organization.organizationName, defaults={"is_active": True}
            )

        # Get employer from employee profile
        if not employer:
            employee_profile = Employee.objects.filter(user=user).first()
            if employee_profile:
                employer = employee_profile.employer

        # Staff fallback
        if not employer and user.is_staff:
            employer = Employer.objects.first()

        if not employer:
            return Response(
                {
                    "error": "You must be associated with an organization to send invitations"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = self.get_serializer(
            data=request.data, context={"employer": employer, "user": user}
        )
        serializer.is_valid(raise_exception=True)
        invitation = serializer.save()

        # SEND OTP EMAIL with HTML template
        
        try:
            otp_expiry = invitation.otp_expires_at.strftime(
                '%B %d, %Y at %I:%M %p'
            )
            subject = f"Welcome to {employer.name} on Obeeoma!"

            # Render HTML template
            html_content = render_to_string('emails/org.html', {
                'employer_name': employer.name,
                'employee_name': "Team Member",  # No employeeName field in model
                'invitation_code': invitation.otp,
                'expiry_date': otp_expiry,
                'department': invitation.employeedepartment,
                'position': None,  # No position field in model
                'message': invitation.message
            })
            
            # Plain text version for fallback
            message = f"""
Hello,

You have been invited to join {employer.name} on Obeeoma by {user.username}.

{invitation.message or ""}

Your 6-digit verification code is: {invitation.otp}

This code expires on {otp_expiry} (valid for 7 days).

If you did not expect this invitation, please ignore this email.

Best regards,
The Obeeoma Team
"""
            
            # Send email with detailed logging
            logger.info(f"Attempting to send invitation email to: {invitation.email}")
            logger.info(f"Subject: {subject}")
            logger.info(f"Employer: {employer.name}")
            logger.info(f"Invitation code: {invitation.otp}")
            
            # Try HTML email first
            success = send_gmail_api_email(invitation.email, subject, message, html_body=html_content)
            
            if success:
                logger.info(f"Invitation email sent successfully to {invitation.email}")
            else:
                logger.error(f"HTML email failed, trying plain text to {invitation.email}")
                # Fallback to plain text only
                success = send_gmail_api_email(invitation.email, subject, message)
                if success:
                    logger.info(f"Plain text invitation email sent to {invitation.email}")
                else:
                    logger.error(f"Both HTML and plain text email failed to {invitation.email}")
                
        except Exception as e:
            logger.exception(f"Failed to send invitation email to {invitation.email}: {str(e)}")

        return Response(
            EmployeeInvitationResponseSerializer(invitation).data,
            status=status.HTTP_201_CREATED,
        )

    # UPDATE INVITATION
    @extend_schema(
        request=EmployeeInvitationCreateSerializer,
        responses={200: EmployeeInvitationResponseSerializer},
        description="Update an existing invitation (message, phone, department)"
    )
    def update(self, request, *args, **kwargs):
        """Update invitation details"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Check if invitation is already accepted
        if instance.accepted:
            return Response(
                {"error": "Cannot update an accepted invitation"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        return Response(
            EmployeeInvitationResponseSerializer(serializer.instance).data,
            status=status.HTTP_200_OK
        )

    # PARTIAL UPDATE
    @extend_schema(
        request=EmployeeInvitationCreateSerializer,
        responses={200: EmployeeInvitationResponseSerializer},
        description="Partially update an invitation"
    )
    def partial_update(self, request, *args, **kwargs):
        """Partially update invitation"""
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    # DELETE INVITATION
    @extend_schema(
        responses={204: None, 400: {"error": "string"}},
        description="Delete/cancel an invitation"
    )
    def destroy(self, request, *args, **kwargs):
        """Delete/cancel an invitation"""
        instance = self.get_object()
        
        # Check if invitation is already accepted
        if instance.accepted:
            return Response(
                {"error": "Cannot delete an accepted invitation"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        self.perform_destroy(instance)
        return Response(
            {"message": "Invitation deleted successfully"},
            status=status.HTTP_200_OK
        )


# ===== ASSESSMENT QUESTIONNAIRE VIEWS =====


@extend_schema_view(
    list=extend_schema(tags=["Assessments - Questions"]),
    retrieve=extend_schema(tags=["Assessments - Questions"]),
)
class AssessmentQuestionViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for assessment questions"""

    queryset = AssessmentQuestion.objects.filter(is_active=True)
    serializer_class = AssessmentQuestionSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["assessment_type"]

    @extend_schema(
        description="Get all questions for a specific assessment type (PHQ-9 or GAD-7)",
        parameters=[
            OpenApiParameter(
                name="type",
                type=str,
                enum=["PHQ-9", "GAD-7"],
                required=True,
                description="Assessment type",
            )
        ],
        tags=["Assessments - Questions"],
    )
    @action(detail=False, methods=["get"])
    def by_type(self, request):
        """Get all questions for a specific assessment with full details"""
        assessment_type = request.query_params.get("type", "PHQ-9")

        if assessment_type not in ["PHQ-9", "GAD-7"]:
            return Response(
                {"error": "Invalid assessment type. Must be PHQ-9 or GAD-7"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        questions = AssessmentQuestion.objects.filter(
            assessment_type=assessment_type, is_active=True
        ).order_by("question_number")

        # Prepare response data
        if assessment_type == "PHQ-9":
            data = {
                "assessment_type": "PHQ-9",
                "title": "Patient Health Questionnaire (PHQ-9)",
                "description": "A 9-question screening tool for depression",
                "instructions": "Over the last 2 weeks, how often have you been bothered by any of the following problems?",
                "time_frame": "Last 2 weeks",
                "questions": AssessmentQuestionSerializer(questions, many=True).data,
                "score_options": [
                    {"value": 0, "label": "Not at all"},
                    {"value": 1, "label": "Several days"},
                    {"value": 2, "label": "More than half the days"},
                    {"value": 3, "label": "Nearly every day"},
                ],
            }
        else:  # GAD-7
            data = {
                "assessment_type": "GAD-7",
                "title": "Generalized Anxiety Disorder (GAD-7)",
                "description": "A 7-question screening tool for anxiety",
                "instructions": "Over the last 2 weeks, how often have you been bothered by any of the following problems?",
                "time_frame": "Last 2 weeks",
                "questions": AssessmentQuestionSerializer(questions, many=True).data,
                "score_options": [
                    {"value": 0, "label": "Not at all"},
                    {"value": 1, "label": "Several days"},
                    {"value": 2, "label": "Over half the days"},
                    {"value": 3, "label": "Nearly every day"},
                ],
            }

        return Response(data)


@extend_schema_view(
    list=extend_schema(tags=["Assessments - Responses"]),
    create=extend_schema(tags=["Assessments - Responses"]),
    retrieve=extend_schema(tags=["Assessments - Responses"]),
)
class AssessmentResponseViewSet(viewsets.ModelViewSet):
    """ViewSet for assessment responses"""

    serializer_class = AssessmentResponseSerializer
    permission_classes = [permissions.IsAuthenticated]
    http_method_names = ["get", "post"]

    def get_queryset(self):
        return AssessmentResponse.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        # Track assessment feature usage
        from .utils.feature import FeatureUsageCalculator
        FeatureUsageCalculator.track_feature(self.request.user, 'assessment')

    @extend_schema(
        description="Get user's assessment history",
        parameters=[
            OpenApiParameter(
                name="type",
                type=str,
                enum=["PHQ-9", "GAD-7"],
                description="Filter by assessment type",
            )
        ],
        responses=AssessmentResponseSerializer(many=True),
        tags=["Assessments - Responses"],
    )
    @action(detail=False, methods=["get"])
    def history(self, request):
        """Get user's assessment history"""
        queryset = self.get_queryset()

        assessment_type = request.query_params.get("type")
        if assessment_type:
            queryset = queryset.filter(assessment_type=assessment_type)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @extend_schema(
        description="Get latest assessment result",
        parameters=[
            OpenApiParameter(
                name="type",
                type=str,
                enum=["PHQ-9", "GAD-7"],
                required=True,
                description="Assessment type",
            )
        ],
        responses=AssessmentResponseSerializer,
        tags=["Assessments - Responses"],
    )
    @action(detail=False, methods=["get"])
    def latest(self, request):
        """Get user's latest assessment result"""
        assessment_type = request.query_params.get("type")

        if not assessment_type:
            return Response(
                {"error": "Assessment type is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            latest = (
                self.get_queryset()
                .filter(assessment_type=assessment_type)
                .latest("completed_at")
            )

            serializer = self.get_serializer(latest)
            return Response(serializer.data)

        except AssessmentResponse.DoesNotExist:
            return Response(
                {"message": f"No {assessment_type} assessment found"},
                status=status.HTTP_404_NOT_FOUND,
            )


# --- Employer Dashboard ---
@extend_schema(tags=["Employer Dashboard"])
class OverviewView(viewsets.ViewSet):
    # permission_classes = [IsCompanyAdmin]

    def list(self, request):
        employee_count = Employee.objects.count()
        active_subscriptions = Subscription.objects.filter(is_active=True).count()
        recent = RecentActivity.objects.select_related("employer").order_by(
            "-timestamp"
        )[:10]
        recent_serialized = RecentActivitySerializer(recent, many=True).data
        return Response(
            {
                "employee_count": employee_count,
                "active_subscriptions": active_subscriptions,
                "recent_activities": recent_serialized,
            }
        )


# Employer Dashboard - Mood Trends (aggregated daily moods for visualization)
@extend_schema(tags=["Employer Dashboard"])
class TrendsView(viewsets.ReadOnlyModelViewSet):
    serializer_class = None  # returning computed data, not model serializer
    permission_classes = [IsCompanyAdmin]

    def list(self, request, *args, **kwargs):
        """
        GET /dashboard/trends/?days=7 or ?start=YYYY-MM-DD&end=YYYY-MM-DD

        Returns an array of day buckets:
        [
          { "date": "2025-02-01", "avg_score": 4.0, "mood_counts": {"Happy": 3, "Neutral": 1} },
          ...
        ]

        Notes:
        - If no date params are provided, all records are used (continuous).
        - Frontend can map avg_score to emojis; we only provide numeric scores and counts.
        - If you need employer scoping, add a filter linking MoodTracking -> Employer.
        """
        qs = MoodTracking.objects.select_related("employee", "user")

        # Date filtering
        start = request.query_params.get("start")
        end = request.query_params.get("end")
        days = request.query_params.get("days")
        today = timezone.now().date()

        if start:
            qs = qs.filter(checked_in_at__date__gte=start)
        if end:
            qs = qs.filter(checked_in_at__date__lte=end)
        if not start and not end and days:
            try:
                days_int = int(days)
                qs = qs.filter(
                    checked_in_at__date__gte=today - timedelta(days=days_int)
                )
            except ValueError:
                pass  # ignore bad days param

        # Aggregate per day
        trend = {}
        for entry in qs.values("checked_in_at__date", "mood"):
            day = entry["checked_in_at__date"]
            mood = entry["mood"] or "Unknown"
            score = MOOD_SCORES.get(mood, 0)

            bucket = trend.setdefault(
                day,
                {
                    "date": day.isoformat(),
                    "total_score": 0,
                    "count": 0,
                    "mood_counts": {},
                },
            )
            bucket["total_score"] += score
            bucket["count"] += 1
            bucket["mood_counts"][mood] = bucket["mood_counts"].get(mood, 0) + 1

        data = []
        for day in sorted(trend.keys()):
            bucket = trend[day]
            avg_score = (
                round(bucket["total_score"] / bucket["count"], 2)
                if bucket["count"]
                else 0
            )
            data.append(
                {
                    "date": bucket["date"],
                    "avg_score": avg_score,
                    "mood_counts": bucket["mood_counts"],
                }
            )

        return Response(data)


@extend_schema(tags=["Employer Dashboard"])
class EmployeeEngagementView(viewsets.ModelViewSet):
    queryset = EmployeeEngagement.objects.select_related("employer").order_by("-month")
    serializer_class = EmployeeEngagementSerializer


from reportlab.pdfgen import canvas

# REPORT 1: DEPARTMENT ANALYSIS REPORT


class DepartmentAnalysisReportView(APIView):
    """
    Generates a downloadable PDF containing department analysis.

    Binary download means:
    - Response is not JSON
    - It is a PDF file returned as bytes
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Create temporary buffer in memory
        buffer = io.BytesIO()

        # Generate PDF using reportlab
        p = canvas.Canvas(buffer)

        p.drawString(100, 800, "DEPARTMENT ANALYSIS REPORT")
        p.drawString(100, 780, f"Requested by: {request.user.username}")
        p.drawString(100, 760, "-----------------------------------------")
        p.drawString(100, 740, "Department statistics go here...")

        p.showPage()
        p.save()

        buffer.seek(0)

        # Response as downloadable PDF
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = (
            'attachment; filename="department-analysis.pdf"'
        )

        return response


#############################################
# REPORT 2: RISK ASSESSMENT REPORT
#############################################


class RiskAssessmentReportView(APIView):
    """
    Downloads a risk assessment report in PDF format.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)

        p.drawString(100, 800, "RISK ASSESSMENT REPORT")
        p.drawString(100, 780, f"Requested by: {request.user.username}")
        p.drawString(100, 760, "-----------------------------------------")
        p.drawString(100, 740, "Risk metrics go here...")

        p.showPage()
        p.save()
        buffer.seek(0)

        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="risk-assessment.pdf"'

        return response


#############################################
# REPORT 3: ENGAGEMENT REPORT
#############################################


class EngagementReportView(APIView):
    """
    Downloads an employee engagement report as PDF.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)

        p.drawString(100, 800, "ENGAGEMENT REPORT")
        p.drawString(100, 780, f"Requested by: {request.user.username}")
        p.drawString(100, 760, "-----------------------------------------")
        p.drawString(100, 740, "Employee engagement trends go here...")

        p.showPage()
        p.save()
        buffer.seek(0)

        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="engagement.pdf"'

        return response


@extend_schema_view(
    list=extend_schema(operation_id="features_usage_list", tags=["Employer Dashboard"]),
    by_category=extend_schema(
        operation_id="features_usage_by_category",
        tags=["Employer Dashboard"],
        description="Returns all active employees with count.",
    ),
)
class EmployeeDashboardView(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer

    @action(detail=False, methods=["get"])
    def active(self, request):
        employees = Employee.objects.filter(status="active")
        serializer = EmployeeSerializer(employees, many=True)
        return Response({"count": employees.count(), "employees": serializer.data})

    @extend_schema(
        operation_id="inactive_employees",
        tags=["Employer Dashboard"],
        description="Returns all inactive employees with count.",
    )
    @action(detail=False, methods=["get"])
    def inactive(self, request):
        employees = Employee.objects.filter(status="inactive")
        serializer = EmployeeSerializer(employees, many=True)
        return Response({"count": employees.count(), "employees": serializer.data})

    @extend_schema(
        operation_id="employee_summary",
        tags=["Employer Dashboard"],
        description="Returns summary of active/inactive employees with percentages.",
    )
    @action(detail=False, methods=["get"])
    def summary(self, request):
        total = Employee.objects.count()
        active = Employee.objects.filter(status="active").count()
        inactive = Employee.objects.filter(status="inactive").count()

        return Response(
            {
                "total": total,
                "active": active,
                "inactive": inactive,
                "active_percent": round((active / total) * 100, 2) if total else 0,
                "inactive_percent": round((inactive / total) * 100, 2) if total else 0,
            }
        )


class FeaturesUsageView(viewsets.ModelViewSet):
    queryset = AIManagement.objects.select_related("employer").order_by("-created_at")
    serializer_class = AIManagementSerializer
    # permission_classes = [IsCompanyAdmin]

    @action(detail=False, methods=["get"])
    def by_category(self, request):
        return Response({"message": "Feature flags grouped by category"})


@extend_schema(tags=["Employer Dashboard"])
class BillingView(viewsets.ModelViewSet):
    queryset = Subscription.objects.select_related("employer").all()
    serializer_class = SubscriptionSerializer
    # permission_classes = [IsCompanyAdmin]

    @action(detail=False, methods=["get"])
    def summary(self, request):
        subscriptions = self.get_queryset()
        total_revenue = sum(float(s.amount) for s in subscriptions)
        return Response(
            {
                "subscriptions": SubscriptionSerializer(subscriptions, many=True).data,
                "total_revenue": total_revenue,
            }
        )


@extend_schema(tags=["Employer Dashboard"])
class UsersView(viewsets.ModelViewSet):
    queryset = Employee.objects.select_related("employer").all()
    serializer_class = EmployeeSerializer
    # permission_classes = [IsCompanyAdmin]


# REPORTS VIEW with DOWNLOAD ACTION
@extend_schema(tags=["Employer Dashboard"])
class ReportsView(viewsets.ReadOnlyModelViewSet):
    queryset = RecentActivity.objects.select_related("employer").order_by("-timestamp")
    serializer_class = RecentActivitySerializer

    # This is the New action for downloading a report
    @action(detail=True, methods=["get"], url_path="download")
    def download_report(self, request, pk=None):
        # Get the report instance
        report = self.get_object()

        # Assuming your model has a file_path field (update if different)
        file_path = getattr(report, "file_path", None)
        if not file_path:
            return Response({"detail": "Report file not found."}, status=404)

        response = FileResponse(open(file_path, "rb"))
        response["Content-Disposition"] = (
            f'attachment; filename="{getattr(report, "title", "report")}.pdf"'
        )
        return response


# For crisis insights about hotline activities and for which reasons employees are reaching out.
@extend_schema(tags=["Employer Dashboard"])
class CrisisInsightsView(viewsets.ReadOnlyModelViewSet):
    queryset = HotlineActivity.objects.select_related("employer").order_by(
        "-recorded_at"
    )
    serializer_class = HotlineActivitySerializer
    permission_classes = [IsCompanyAdmin]


def home(request):
    return JsonResponse({"status": "ok", "app": "obeeomaapp"})


@extend_schema(tags=["Authentication"])
class EmailConfigCheckView(APIView):
    """Debug endpoint to check email configuration"""

    permission_classes = [permissions.AllowAny]

    def get(self, request):
        import os

        db_config = settings.DATABASES["default"]

        config = {
            "database": {
                "engine": db_config.get("ENGINE"),
                "name": str(db_config.get("NAME")),
                "host": db_config.get("HOST", "Not set"),
                "user": db_config.get("USER", "Not set"),
            },
            "environment_vars": {
                "DATABASE_URL": "Set" if os.getenv("DATABASE_URL") else "NOT SET",
                "PGHOST": "Set" if os.getenv("PGHOST") else "NOT SET",
            },
            "email": {
                "email_backend": settings.EMAIL_BACKEND,
                "email_host": settings.EMAIL_HOST,
                "email_port": settings.EMAIL_PORT,
                "email_use_tls": settings.EMAIL_USE_TLS,
                "email_host_user": settings.EMAIL_HOST_USER,
                "default_from_email": settings.DEFAULT_FROM_EMAIL,
                "has_email_password": bool(settings.EMAIL_HOST_PASSWORD),
            },
            "debug_mode": settings.DEBUG,
        }
        return Response(config)




# EmployeeProfileView
@extend_schema(tags=["Employee - Profile"])
class EmployeeProfileView(viewsets.ModelViewSet):
    """
    Employee can view/update their own profile.
    """
    serializer_class = EmployeeProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user.employee_public_profile

    def get_queryset(self):
        # Only return current user's profile
        return EmployeeProfile.objects.filter(user=self.request.user)

    def update(self, request, *args, **kwargs):
        """Handle PUT requests to update profile"""
        partial = False
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        """Handle PATCH requests to update profile"""
        partial = True
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


# AvatarProfileView
@extend_schema(tags=["Employee - Profile"])
class AvatarProfileView(viewsets.ModelViewSet):
    serializer_class = AvatarProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return AvatarProfile.objects.filter(employee__user=self.request.user)


#mood tracking view 
@extend_schema(tags=['Employee - Mood Tracking'])
class MoodTrackingView(viewsets.ModelViewSet):
    serializer_class = MoodTrackingSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ["mood"]
    search_fields = ["note"]

    def get_queryset(self):
        # Employees only see their own data
        return MoodTracking.objects.filter(employee__user=self.request.user)

    def perform_create(self, serializer):
        employee = get_object_or_404(EmployeeProfile, user=self.request.user)
        mood_entry = serializer.save(user=self.request.user, employee=employee)
        
        # Auto-sync to wellness graph
        self._sync_to_wellness_graph(self.request.user, mood_entry)
        
        return mood_entry

   
    # EMPLOYEE: Weekly Mood Summary
    @action(detail=False, methods=['get'], url_path='mood-summary')
    def mood_summary(self, request):
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        today = now().date()
        start_date = today - timedelta(days=6)

        # Pre-fill all days as "Missed"
        week_days = [(start_date + timedelta(days=i)) for i in range(7)]
        summary = {day.strftime('%A'): "Missed" for day in week_days}

        checkins = MoodTracking.objects.filter(
            employee=employee,
            checked_in_at__date__range=(start_date, today)
        )

        for checkin in checkins:
            day_name = checkin.checked_in_at.strftime('%A')
            summary[day_name] = checkin.mood

        return Response(summary)

    # ============================
    # EMPLOYER: Company Mood Summary
    # ============================
    @extend_schema(tags=['Employer - Mood Tracking'])
    @action(
        detail=False,
        methods=['get'],
        url_path='employer-summary',
        permission_classes=[IsAuthenticated]
    )
    def employer_mood_summary(self, request):
        today = now().date()
        start_date = today - timedelta(days=6)

        qs = MoodTracking.objects.filter(
            checked_in_at__date__range=(start_date, today)
        )

        overall = qs.aggregate(
            average_mood=Avg('mood'),
            total_entries=Count('id')
        )

        distribution = (
            qs.values('mood')
            .annotate(count=Count('mood'))
            .order_by('mood')
        )

        daily_avg = (
            qs.values('checked_in_at__date')
            .annotate(avg_mood=Avg('mood'))
            .order_by('checked_in_at__date')
        )

        return Response({
            "period": "last_7_days",
            "average_mood": round(overall['average_mood'], 2) if overall['average_mood'] else 0,
            "total_entries": overall['total_entries'],
            "mood_distribution": distribution,
            "daily_average": daily_avg
        })

    # ============================
    # MOOD TRACKING SCREEN
    # ============================
    @action(detail=False, methods=['get'], url_path='screen')
    def screen(self, request):
        """Get mood tracking screen overview"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        today = now().date()
        
        # Check if user already checked in today
        today_checkin = MoodTracking.objects.filter(
            employee=employee,
            checked_in_at__date=today
        ).first()
        
        # Get recent mood entries (last 7 days)
        recent_entries = MoodTracking.objects.filter(
            employee=employee,
            checked_in_at__date__gte=today - timedelta(days=7)
        ).order_by('-checked_in_at')
        
        # Calculate mood streak
        streak = self._calculate_mood_streak(employee)
        
        return Response({
            "has_checked_in_today": bool(today_checkin),
            "today_mood": today_checkin.mood if today_checkin else None,
            "current_streak": streak,
            "recent_entries_count": recent_entries.count(),
            "mood_categories": MoodTracking.MOOD_CATEGORIES
        })

    # ============================
    # MOOD ENTRIES
    # ============================
    @action(detail=False, methods=['get', 'post'], url_path='entries')
    def entries(self, request):
        """Get or create mood entries"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        
        if request.method == 'GET':
            # Get mood entries with optional date filtering
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            queryset = MoodTracking.objects.filter(employee=employee)
            
            if start_date:
                queryset = queryset.filter(checked_in_at__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(checked_in_at__date__lte=end_date)
                
            entries = queryset.order_by('-checked_in_at')
            serializer = self.get_serializer(entries, many=True)
            return Response(serializer.data)
            
        elif request.method == 'POST':
            # Create new mood entry
            mood = request.data.get('mood')
            note = request.data.get('note', '')
            
            if not mood:
                return Response(
                    {"error": "Mood is required"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if already checked in today
            today = now().date()
            existing = MoodTracking.objects.filter(
                employee=employee,
                checked_in_at__date=today
            ).first()
            
            if existing:
                return Response(
                    {"error": "Already checked in today"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            mood_entry = MoodTracking.objects.create(
                user=request.user,
                employee=employee,
                mood=mood,
                note=note
            )
            
            serializer = self.get_serializer(mood_entry)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    # ============================
    # TODAY'S MOOD
    # ============================
    @action(detail=False, methods=['get', 'put'], url_path='today')
    def today(self, request):
        """Get or update today's mood"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        today = now().date()
        
        mood_entry = MoodTracking.objects.filter(
            employee=employee,
            checked_in_at__date=today
        ).first()
        
        if request.method == 'GET':
            if mood_entry:
                serializer = self.get_serializer(mood_entry)
                return Response(serializer.data)
            else:
                return Response({"message": "No mood entry for today"})
        
        elif request.method == 'PUT':
            if not mood_entry:
                return Response(
                    {"error": "No mood entry found for today"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            mood = request.data.get('mood')
            note = request.data.get('note', '')
            
            if mood:
                mood_entry.mood = mood
            if note:
                mood_entry.note = note
            
            mood_entry.save()
            serializer = self.get_serializer(mood_entry)
            return Response(serializer.data)

    # ============================
    # MOOD CHART
    # ============================
    @action(detail=False, methods=['get'], url_path='chart')
    def chart(self, request):
        """Get mood data for chart visualization"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        
        # Get period from query params (default: 30 days)
        period = int(request.query_params.get('period', 30))
        start_date = now().date() - timedelta(days=period)
        
        entries = MoodTracking.objects.filter(
            employee=employee,
            checked_in_at__date__gte=start_date
        ).order_by('checked_in_at')
        
        # Convert mood to numeric values for charting
        mood_values = {
            'Ecstatic': 5, 'Happy': 4, 'Excited': 4, 'Content': 3,
            'Calm': 3, 'Neutral': 2, 'Tired': 2,
            'Anxious': 1, 'Stressed': 1, 'Sad': 1, 'Frustrated': 1, 'Angry': 0
        }
        
        chart_data = []
        for entry in entries:
            chart_data.append({
                'date': entry.checked_in_at.strftime('%Y-%m-%d'),
                'mood': entry.mood,
                'value': mood_values.get(entry.mood, 2),
                'note': entry.note
            })
        
        return Response({
            "period": f"last_{period}_days",
            "data": chart_data,
            "average_mood": sum(item['value'] for item in chart_data) / len(chart_data) if chart_data else 0
        })

    # ============================
    # MOOD ANALYTICS
    # ============================
    @action(detail=False, methods=['get'], url_path='analytics')
    def analytics(self, request):
        """Get comprehensive mood analytics"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        
        # Get period from query params (default: 30 days)
        period = int(request.query_params.get('period', 30))
        start_date = now().date() - timedelta(days=period)
        
        entries = MoodTracking.objects.filter(
            employee=employee,
            checked_in_at__date__gte=start_date
        )
        
        # Mood distribution
        mood_distribution = (
            entries.values('mood')
            .annotate(count=Count('mood'))
            .order_by('-count')
        )
        
        # Category distribution
        category_stats = {}
        for mood, category in MoodTracking.MOOD_CATEGORIES.items():
            category_stats[category] = entries.filter(mood=mood).count()
        
        # Weekly patterns
        weekly_pattern = {}
        for i in range(7):
            day_entries = entries.filter(
                checked_in_at__week_day=(i + 1) % 7  # Django uses 0-6 for Sunday-Saturday
            )
            if day_entries.exists():
                avg_mood = day_entries.aggregate(avg=Avg('mood'))['avg']
                weekly_pattern[calendar.day_name[i]] = {
                    'count': day_entries.count(),
                    'most_common': day_entries.values('mood').annotate(
                        count=Count('mood')
                    ).order_by('-count').first()['mood'] if day_entries.exists() else None
                }
        
        # Streak information
        current_streak = self._calculate_mood_streak(employee)
        longest_streak = self._calculate_longest_streak(employee)
        
        return Response({
            "period": f"last_{period}_days",
            "total_entries": entries.count(),
            "mood_distribution": list(mood_distribution),
            "category_distribution": category_stats,
            "weekly_pattern": weekly_pattern,
            "current_streak": current_streak,
            "longest_streak": longest_streak,
            "check_in_rate": round((entries.count() / period) * 100, 1)
        })

    # ============================
    # MOOD HISTORY
    # ============================
    @action(detail=False, methods=['get'], url_path='history')
    def history(self, request):
        """Get mood history with pagination"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        
        # Pagination parameters
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        queryset = MoodTracking.objects.filter(employee=employee)
        
        if start_date:
            queryset = queryset.filter(checked_in_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(checked_in_at__date__lte=end_date)
        
        # Order by date descending
        queryset = queryset.order_by('-checked_in_at')
        
        # Manual pagination
        total_count = queryset.count()
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        
        entries = queryset[start_idx:end_idx]
        serializer = self.get_serializer(entries, many=True)
        
        return Response({
            "entries": serializer.data,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total_count": total_count,
                "total_pages": (total_count + page_size - 1) // page_size
            }
        })

    # ============================
    # MOOD PATTERNS
    # ============================
    @action(detail=False, methods=['get'], url_path='patterns')
    def patterns(self, request):
        """Analyze mood patterns and trends"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        
        # Get period from query params (default: 90 days for pattern analysis)
        period = int(request.query_params.get('period', 90))
        start_date = now().date() - timedelta(days=period)
        
        entries = MoodTracking.objects.filter(
            employee=employee,
            checked_in_at__date__gte=start_date
        ).order_by('checked_in_at')
        
        if entries.count() < 7:
            return Response({
                "error": "Insufficient data for pattern analysis. Need at least 7 entries.",
                "patterns": {}
            })
        
        patterns = {}
        
        # Time of day patterns
        time_patterns = {'morning': [], 'afternoon': [], 'evening': []}
        for entry in entries:
            hour = entry.checked_in_at.hour
            if 6 <= hour < 12:
                time_patterns['morning'].append(entry.mood)
            elif 12 <= hour < 18:
                time_patterns['afternoon'].append(entry.mood)
            else:
                time_patterns['evening'].append(entry.mood)
        
        for time_period, moods in time_patterns.items():
            if moods:
                most_common = max(set(moods), key=moods.count)
                patterns[f'{time_period}_most_common'] = most_common
                patterns[f'{time_period}_entry_count'] = len(moods)
        
        # Day of week patterns
        day_patterns = {}
        for entry in entries:
            day_name = entry.checked_in_at.strftime('%A')
            if day_name not in day_patterns:
                day_patterns[day_name] = []
            day_patterns[day_name].append(entry.mood)
        
        for day, moods in day_patterns.items():
            if moods:
                most_common = max(set(moods), key=moods.count)
                patterns[f'{day.lower()}_most_common'] = most_common
        
        # Mood sequences (consecutive days)
        sequences = self._analyze_mood_sequences(entries)
        patterns['mood_sequences'] = sequences
        
        # Improvement indicators
        recent_moods = entries.order_by('-checked_in_at')[:14]  # Last 2 weeks
        older_moods = entries.order_by('-checked_in_at')[14:28]  # Previous 2 weeks
        
        if recent_moods.count() >= 7 and older_moods.count() >= 7:
            recent_positive = sum(1 for m in recent_moods if MoodTracking.MOOD_CATEGORIES.get(m.mood) == 'Positive')
            older_positive = sum(1 for m in older_moods if MoodTracking.MOOD_CATEGORIES.get(m.mood) == 'Positive')
            
            patterns['trend'] = 'improving' if recent_positive > older_positive else 'declining' if recent_positive < older_positive else 'stable'
        
        return Response({"patterns": patterns})

    # ============================
    # MOOD INSIGHTS
    # ============================
    @action(detail=False, methods=['get'], url_path='insights')
    def insights(self, request):
        """Get personalized mood insights"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        
        # Get period from query params (default: 30 days)
        period = int(request.query_params.get('period', 30))
        start_date = now().date() - timedelta(days=period)
        
        entries = MoodTracking.objects.filter(
            employee=employee,
            checked_in_at__date__gte=start_date
        )
        
        insights = []
        
        # Generate insights based on data
        if entries.count() >= 5:
            # Most common mood
            most_common = entries.values('mood').annotate(count=Count('mood')).order_by('-count').first()
            if most_common:
                insights.append({
                    "type": "most_common_mood",
                    "title": f"Your most common mood is {most_common['mood']}",
                    "description": f"You've logged this mood {most_common['count']} times in the last {period} days.",
                    "priority": "medium"
                })
            
            # Check-in consistency
            check_in_rate = (entries.count() / period) * 100
            if check_in_rate >= 80:
                insights.append({
                    "type": "consistency",
                    "title": "Great consistency!",
                    "description": f"You've checked in {check_in_rate:.1f}% of days. Keep it up!",
                    "priority": "positive"
                })
            elif check_in_rate < 50:
                insights.append({
                    "type": "consistency",
                    "title": "Try to be more consistent",
                    "description": f"You've only checked in {check_in_rate:.1f}% of days. Regular check-ins help track your mood better.",
                    "priority": "high"
                })
            
            # Mood balance
            positive_count = sum(1 for e in entries if MoodTracking.MOOD_CATEGORIES.get(e.mood) == 'Positive')
            negative_count = sum(1 for e in entries if MoodTracking.MOOD_CATEGORIES.get(e.mood) == 'Negative')
            
            if negative_count > positive_count * 1.5:
                insights.append({
                    "type": "mood_balance",
                    "title": "More negative moods detected",
                    "description": "You've been logging more negative moods lately. Consider trying stress-reduction techniques.",
                    "priority": "high"
                })
        
        return Response({
            "insights": insights,
            "generated_at": now().isoformat(),
            "period": f"last_{period}_days"
        })

    # ============================
    # UNREAD INSIGHTS
    # ============================
    @action(detail=False, methods=['get'], url_path='unread-insights')
    def unread_insights(self, request):
        """Get unread mood insights"""
        # This would typically integrate with a notification/read status system
        # For now, return recent insights that haven't been "read"
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        
        # Get last time insights were viewed (simplified - using cache)
        cache_key = f'mood_insights_viewed_{employee.id}'
        last_viewed = cache.get(cache_key)
        
        # Get recent insights
        insights_response = self.insights(request)
        insights = insights_response.data.get('insights', [])
        
        # Filter for "unread" insights (created after last view)
        if last_viewed:
            # In a real implementation, you'd filter by creation timestamp
            # For now, return all insights as unread
            pass
        
        return Response({
            "unread_count": len(insights),
            "insights": insights
        })

    # ============================
    # MARK INSIGHTS AS READ
    # ============================
    @action(detail=False, methods=['post'], url_path='mark-read')
    def mark_read(self, request):
        """Mark insights as read"""
        employee = get_object_or_404(EmployeeProfile, user=request.user)
        
        # Mark current time as last viewed
        cache_key = f'mood_insights_viewed_{employee.id}'
        cache.set(cache_key, now().isoformat(), timeout=86400)  # Cache for 24 hours
        
        return Response({
            "message": "Insights marked as read",
            "timestamp": now().isoformat()
        })

    # ============================
    # HELPER METHODS
    # ============================
    def _sync_to_wellness_graph(self, user, mood_entry):
        """Automatically sync mood entry to wellness graph"""
        mood_scores = {
            'Ecstatic': 5, 'Happy': 4, 'Excited': 4, 'Content': 3,
            'Calm': 3, 'Neutral': 2, 'Tired': 2,
            'Anxious': 1, 'Stressed': 1, 'Sad': 1, 'Frustrated': 1, 'Angry': 0
        }
        
        mood_date = mood_entry.checked_in_at.date()
        mood_score = mood_scores.get(mood_entry.mood, 2)
        
        # Create or update wellness graph entry
        WellnessGraph.objects.update_or_create(
            user=user,
            mood_date=mood_date,
            defaults={'mood_score': mood_score}
        )

    def _calculate_mood_streak(self, employee):
        """Calculate current mood check-in streak"""
        today = now().date()
        streak = 0
        
        for i in range(365):  # Check up to a year
            check_date = today - timedelta(days=i)
            has_checkin = MoodTracking.objects.filter(
                employee=employee,
                checked_in_at__date=check_date
            ).exists()
            
            if has_checkin:
                streak += 1
            else:
                break
        
        return streak

    def _calculate_longest_streak(self, employee):
        """Calculate longest mood check-in streak"""
        entries = MoodTracking.objects.filter(
            employee=employee
        ).order_by('checked_in_at')
        
        if not entries.exists():
            return 0
        
        longest_streak = 0
        current_streak = 1
        last_date = entries.first().checked_in_at.date()
        
        for entry in entries[1:]:
            current_date = entry.checked_in_at.date()
            if (current_date - last_date).days == 1:
                current_streak += 1
            else:
                longest_streak = max(longest_streak, current_streak)
                current_streak = 1
            last_date = current_date
        
        return max(longest_streak, current_streak)

    def _analyze_mood_sequences(self, entries):
        """Analyze consecutive mood patterns"""
        if not entries.exists():
            return []
        
        sequences = []
        current_sequence = [entries.first().mood]
        
        for entry in entries[1:]:
            if entry.mood == current_sequence[-1]:
                current_sequence.append(entry.mood)
            else:
                if len(current_sequence) >= 3:  # Only consider sequences of 3+ days
                    sequences.append({
                        "mood": current_sequence[0],
                        "length": len(current_sequence),
                        "start_date": entries[0].checked_in_at.strftime('%Y-%m-%d')
                    })
                current_sequence = [entry.mood]
        
        # Check the last sequence
        if len(current_sequence) >= 3:
            sequences.append({
                "mood": current_sequence[0],
                "length": len(current_sequence),
                "start_date": entries[0].checked_in_at.strftime('%Y-%m-%d')
            })
        
        return sequences
@extend_schema(tags=['Employee - Assessments'])
@extend_schema(tags=['Resources'])
class SelfHelpResourceView(viewsets.ModelViewSet):
    queryset = SelfHelpResource.objects.all()
    serializer_class = SelfHelpResourceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return MoodTracking.objects.filter(employee__user=self.request.user)


@extend_schema(tags=["Employee - Crisis Support"])
class CrisisTriggerView(viewsets.ModelViewSet):
    serializer_class = CrisisTriggerSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return CrisisTrigger.objects.filter(employee__user=self.request.user)



# notifications/views.py

@extend_schema(tags=["Employee - Notifications"])
class NotificationView(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["read"]  # Use the actual field name 'read' instead of 'is_read'
    ordering_fields = ["sent_on"]
    ordering = ["-sent_on"]
    filterset_fields = ['read']
    ordering_fields = ['sent_on']
    ordering = ['-sent_on']

    def get_queryset(self):
        return Notification.objects.filter(employee__user=self.request.user)

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.read = True
        notification.save()
        return Response({"status": "read"})

@extend_schema(tags=['Employee - Engagement'])
class EngagementTrackerView(viewsets.ModelViewSet):
    serializer_class = EngagementTrackerSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return EngagementTracker.objects.filter(employee__user=self.request.user)


class FeedbackFilterSet(FilterSet):
    feedback_type = CharFilter(method="filter_feedback_type")

    def filter_feedback_type(self, queryset, name, value):
        return queryset.filter(feedback_type__icontains=value)

    class Meta:
        model = Feedback
        fields = ["rating"]  # Only include actual model fields here


@extend_schema(tags=["Employee - Feedback"])
class FeedbackView(viewsets.ModelViewSet):
    serializer_class = FeedbackSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["rating", "feedback_type"]
    ordering_fields = ["created_at", "rating"]
    ordering = ["-created_at"]

    def get_queryset(self):
        return Feedback.objects.filter(employee__user=self.request.user)

    @action(detail=False, methods=["get"])
    def average_rating(self, request):
        from django.db.models import Avg

        avg = self.get_queryset().aggregate(average=Avg("rating"))
        return Response(
            {
                "average_rating": round(avg["average"], 2) if avg["average"] else 0,
                "total_feedback": self.get_queryset().count(),
            }
        )


@extend_schema(tags=["Employee - AI Chat"])
class ChatSessionView(viewsets.ModelViewSet):
    serializer_class = ChatSessionSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["is_active"]
    ordering_fields = ["started_at", "last_message_at"]
    ordering = ["-started_at"]

    def get_queryset(self):
        return ChatSession.objects.filter(employee__user=self.request.user)

    def perform_create(self, serializer):
        # Get or create EmployeeProfile for the user
        employee, created = EmployeeProfile.objects.get_or_create(
            user=self.request.user,
            defaults={
                'display_name': self.request.user.username,
                'organization': 'Default Organization',
                'role': 'Employee',
            }
        )
        if created:
            print(f"Created EmployeeProfile for user: {self.request.user.email}")
        
        serializer.save(employee=employee)

    @action(detail=False, methods=["get"])
    def active(self, request):
        active_sessions = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(active_sessions, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    def end_session(self, request, pk=None):
        session = self.get_object()
        session.is_active = False
        session.save()
        return Response(
            {
                "message": "Chat session ended successfully",
                "session": self.get_serializer(session).data,
            }
        )


@extend_schema(tags=["Employee - AI Chat"])
class ChatMessageView(viewsets.ModelViewSet):
    serializer_class = ChatMessageSerializer
    permission_classes = [permissions.IsAuthenticated]
    ordering = ["timestamp"]

    def get_queryset(self):
        """
        Restrict messages to the current user's session.
        Ensures employees can only access their own chat sessions.
        """
        return ChatMessage.objects.filter(
            session_id=self.kwargs.get("session_id"),
            session__employee__user=self.request.user,
        )

    def create(self, request, *args, **kwargs):
        """
        Override create to handle the AI chat flow and return AI response
        """
        # Get the chat session for the current user
        session = get_object_or_404(
            ChatSession,
            id=self.kwargs.get("session_id"),
            employee__user=self.request.user,
        )

        # Save the incoming user message
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user_message = serializer.save(session=session, sender="user")
        
        # Track sana_ai feature usage
        from .utils.feature import FeatureUsageCalculator
        FeatureUsageCalculator.track_feature(self.request.user, 'sana_ai')

        # Ensure a system prompt exists in the session (only once per session)
        if not session.messages.filter(sender="system").exists():
            ChatMessage.objects.create(
                session=session,
                sender="system",
                message="You are Sana, a helpful and professional AI assistant.",
            )

        # Build conversation history for Groq API (exclude the current user message)
        conversation_history = []
        for msg in session.messages.all().order_by("timestamp"):
            # Skip the current user message we just saved to avoid duplication
            if msg.id == user_message.id:
                continue
            # Use the model's api_role() helper to map DB roles to Groq roles
            role = msg.api_role()
            conversation_history.append({"role": role, "content": msg.message})

        # Call Groq service to generate AI reply
        try:
            groq_service = GroqService()
            ai_reply = groq_service.get_response(
                user_message=user_message.message,
                conversation_history=conversation_history,
            )

            # Save AI response back into the database
            ai_message = ChatMessage.objects.create(session=session, sender="ai", message=ai_reply)
            
            # Log what we're returning
            print(f"Returning AI response: {ai_reply[:100]}...")
            
            # Return the AI response to the client
            return Response({
                "id": ai_message.id,
                "sender": ai_message.sender,
                "message": ai_message.message,
                "timestamp": ai_message.timestamp
            })

        except ValueError as e:
            # Handle missing API key
            logger.error(f"Groq API configuration error: {str(e)}")
            return Response(
                {"error": "AI service not configured. Please contact support."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        except Exception as e:
            # Log error for debugging
            logger.error(f"Groq chat error: {str(e)}")

            # Return error response to client instead of failing silently
            return Response(
                {"error": "AI service failed. Please try again later."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def perform_create(self, serializer):
        """
        This method is no longer used since we override create() above
        """
        pass


@extend_schema(tags=["Employee - Recommendations"])
class RecommendationLogView(viewsets.ModelViewSet):
    serializer_class = RecommendationLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return RecommendationLog.objects.filter(employee__user=self.request.user)

    def perform_create(self, serializer):
        employee = get_object_or_404(EmployeeProfile, user=self.request.user)
        serializer.save(employee=employee)


# --- Employee-specific: Badges and Engagement Streaks ---
@extend_schema(tags=["Employee - Engagement"])
class MyBadgesView(viewsets.ReadOnlyModelViewSet):
    serializer_class = UserBadgeSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return UserBadge.objects.filter(user=self.request.user)


@extend_schema(tags=["Employee - Engagement"])
class MyStreaksView(viewsets.ReadOnlyModelViewSet):
    serializer_class = EngagementStreakSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return EngagementStreak.objects.filter(user=self.request.user).order_by(
            "-last_active_date"
        )


# --- Employer APIs ---
@extend_schema(tags=["Employer - Organization"])
class EmployerViewSet(viewsets.ModelViewSet):
    queryset = Employer.objects.all()
    serializer_class = EmployerSerializer

    def get_permissions(self):
        """
        Only employers can create organizations (not system admins).
        System admins can only view, update, and delete.
        """
        if self.action == "create":
            # Block creation through this endpoint - use EmployerRegistrationView instead
            return [permissions.IsAuthenticated()]
        elif self.action in ["list", "retrieve"]:
            return [permissions.IsAuthenticated()]
        else:
            # Update, delete require admin
            return [IsCompanyAdmin()]

    def create(self, request, *args, **kwargs):
        """
        Block direct creation - employers should use /auth/register-organization/ endpoint
        """
        return Response(
            {
                "error": "Please use /auth/register-organization/ endpoint to create an organization",
                "detail": "Direct organization creation is not allowed. Use the employer registration endpoint.",
            },
            status=status.HTTP_403_FORBIDDEN,
        )

    def get_queryset(self):
        """
        Employers can only see their own organization.
        Admins can see all organizations.
        """
        if self.request.user.is_staff:
            return Employer.objects.all()
        # Return organizations where user is linked
        return Employer.objects.filter(employees__user=self.request.user).distinct()

    def perform_create(self, serializer):
        """
        When an employer creates an organization, link them to it.
        """
        employer = serializer.save()
        # Create an employee profile linking the user to this organization
        if not hasattr(self.request.user, "employee_record"):
            Employee.objects.create(
                employer=employer,
                user=self.request.user,
                first_name=self.request.user.first_name or self.request.user.username,
                last_name=self.request.user.last_name or "",
                email=self.request.user.email,
                status="active",
            )

    @action(detail=True, methods=["get"])
    def overview(self, request, pk=None):
        employer = self.get_object()
        data = {
            "employee_count": Employee.objects.filter(employer=employer).count(),
            "active_subscriptions": Subscription.objects.filter(
                employer=employer, is_active=True
            ).count(),
            "engagement_entries": EmployeeEngagement.objects.filter(
                employer=employer
            ).count(),
            "latest_hotline_activity": (
                HotlineActivitySerializer(
                    HotlineActivity.objects.filter(employer=employer)
                    .order_by("-recorded_at")
                    .first()
                ).data
                if HotlineActivity.objects.filter(employer=employer).exists()
                else None
            ),
            "recent_activities": RecentActivitySerializer(
                RecentActivity.objects.filter(employer=employer).order_by("-timestamp")[
                    :10
                ],
                many=True,
            ).data,
        }
        return Response(data)

    @action(detail=True, methods=["get"])
    def employees(self, request, pk=None):
        employer = self.get_object()
        qs = Employee.objects.filter(employer=employer).order_by("-joined_date")
        return Response(EmployeeSerializer(qs, many=True).data)

    @action(detail=True, methods=["get"])
    def subscriptions(self, request, pk=None):
        employer = self.get_object()
        qs = Subscription.objects.filter(employer=employer).order_by("-start_date")
        return Response(SubscriptionSerializer(qs, many=True).data)

    @action(detail=True, methods=["get"])
    def features(self, request, pk=None):
        employer = self.get_object()
        qs = AIManagement.objects.filter(employer=employer).order_by("-created_at")
        return Response(AIManagementSerializer(qs, many=True).data)

    @action(detail=True, methods=["get"])
    def engagements(self, request, pk=None):
        employer = self.get_object()
        qs = EmployeeEngagement.objects.filter(employer=employer).order_by("-month")
        return Response(EmployeeEngagementSerializer(qs, many=True).data)

    @action(detail=True, methods=["get"])
    def activities(self, request, pk=None):
        employer = self.get_object()
        qs = RecentActivity.objects.filter(employer=employer).order_by("-timestamp")
        return Response(RecentActivitySerializer(qs, many=True).data)

    @action(detail=True, methods=["post"], permission_classes=[IsCompanyAdmin])
    def invite(self, request, pk=None):
        employer = self.get_object()
        serializer = EmployeeInvitationCreateSerializer(
            data=request.data, context={"employer": employer, "user": request.user}
        )
        serializer.is_valid(raise_exception=True)
        invite = serializer.save()
        return Response(
            {"message": "Invitation created", "token": invite.token},
            status=status.HTTP_201_CREATED,
        )


@extend_schema(tags=["Employee - Progress"])
class ProgressViewSet(viewsets.ModelViewSet):
    queryset = Progress.objects.all()
    serializer_class = ProgressSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAdminUser])
    def analytics(self, request):
        data = {
            "average_mood": Progress.objects.aggregate(Avg("mood_score"))[
                "mood_score__avg"
            ],
            "total_users": User.objects.count(),
            "progress_entries": Progress.objects.count(),
        }
        return Response(data)


# New Dashboard Views
@extend_schema(tags=["Employer Dashboard"])
class OrganizationOverviewView(viewsets.ViewSet):
    """
    EMPLOYER DASHBOARD - ORGANIZATION OVERVIEW

    This view provides a comprehensive overview of the organization's workforce
    and their engagement with the mental health platform.

    Purpose for Employers:
    - Monitor overall workforce size and activity
    - Track employee wellness trends
    - Understand platform feature adoption
    - View recent organizational activities

    Key Metrics Provided:
    - Total employee count
    - Wellness index (aggregate mental health score)
    - Active employees count
    - Feature usage statistics
    - Engagement trends
    """

    permission_classes = [IsCompanyAdmin]

    def list(self, request):
        from django.db.models import Q, Count, Avg
        from datetime import datetime, timedelta

        # ============================================================
        # STEP 1: IDENTIFY THE EMPLOYER/ORGANIZATION
        # ============================================================
        # Get the employer associated with the logged-in user
        # This ensures employers only see their own organization's data
        employer = None
        if hasattr(request.user, "employer_profile"):
            employer = request.user.employer_profile.employer

        # Filter employees by employer if available
        employee_queryset = Employee.objects.all()
        if employer:
            employee_queryset = employee_queryset.filter(employer=employer)

        # ============================================================
        # METRIC 1: TOTAL EMPLOYEES
        # ============================================================
        # Purpose: Shows the total number of employees in the organization
        # Helps employers understand their workforce size
        total_employees = employee_queryset.count()

        # ============================================================
        # METRIC 2: WELLNESS INDEX
        # ============================================================
        # Purpose: Aggregate wellness score across all employees
        # Higher score = better overall mental health
        # Range: 0-100 (100 being best)
        #
        # How it works:
        # - Collects all assessment responses from employees
        # - Calculates average score
        # - Converts to wellness index (inverse of severity)
        #
        # Why it matters for employers:
        # - Quick snapshot of overall workforce mental health
        # - Helps identify if intervention programs are working
        # - Can track improvement over time
        recent_assessments = AssessmentResponse.objects.filter(
            user__in=employee_queryset.values_list("user", flat=True)
        )

        if recent_assessments.exists():
            # Calculate wellness index (inverse of severity - higher is better)
            # Assuming max score is 27 for PHQ-9, convert to percentage
            avg_total_score = (
                recent_assessments.aggregate(avg=Avg("total_score"))["avg"] or 0
            )
            wellness_index = max(0, int(100 - (avg_total_score / 27 * 100)))
        else:
            wellness_index = 0

        # Generate wellness index description based on score
        if wellness_index >= 80:
            wellness_description = (
                "Excellent - Your workforce shows strong mental wellbeing"
            )
        elif wellness_index >= 60:
            wellness_description = "Good - Overall mental health is positive"
        elif wellness_index >= 40:
            wellness_description = "Fair - Some employees may need support"
        elif wellness_index >= 20:
            wellness_description = "Concerning - Consider wellness initiatives"
        else:
            wellness_description = "Critical - Immediate attention recommended"

        # ============================================================
        # METRIC 3: ACTIVE EMPLOYEES COUNT
        # ============================================================
        # Purpose: Shows how many employees are currently active
        # Helps employers understand workforce availability
        # Active status means employee is currently working and not suspended
        active_employees_count = employee_queryset.filter(status="active").count()

        # ============================================================
        # METRIC 4: RECENT EMPLOYEES LIST
        # ============================================================
        # Purpose: Shows the 10 most recently joined employees
        # Helps employers:
        # - Track new hires
        # - Monitor onboarding progress
        # - See department distribution
        employees_list = employee_queryset.select_related("department").order_by(
            "-joined_date"
        )[:10]
        employees_data = [
            {
                "id": emp.id,
                "name": f"{emp.first_name} {emp.last_name}",
                "email": emp.email,
                "department": emp.department.name if emp.department else "N/A",
                "status": emp.status,
            }
            for emp in employees_list
        ]

        # ============================================================
        # METRIC 5: ENGAGEMENT TREND
        # ============================================================
        # Purpose: Shows workforce engagement breakdown
        # Pending = Invited but not yet onboarded
        # Active = Currently working employees
        # Inactive = Employees who are not currently active or suspended
        #
        # Why it matters for employers:
        # - Understand workforce availability
        # - Track onboarding progress (pending employees)
        # - Identify potential retention issues
        # - Plan resource allocation
        engagement_stats = employee_queryset.aggregate(
            pending=Count("id", filter=Q(status="pending")),
            active=Count("id", filter=Q(status="active")),
            inactive=Count("id", filter=Q(status__in=["inactive", "suspended"])),
        )

        # ============================================================
        # METRIC 6: FEATURE USAGE BREAKDOWN
        # ============================================================
        # Purpose: Shows what percentage of employees use each platform feature
        #
        # Features tracked:
        # 1. Wellness Assessments - Mental health self-assessments
        # 2. AI Chatbot (Sana) - AI-powered mental health support
        # 3. Mood Tracking - Daily mood check-ins
        # 4. Resource Library - Educational content and tools
        #
        # Why it matters for employers:
        # - Understand platform adoption
        # - Identify which features are most valuable
        # - Justify ROI on the platform
        # - Identify areas needing more promotion
        total_users = employee_queryset.filter(status="active").count() or 1

        # Count unique users who have used wellness assessments
        wellness_assessments_count = (
            AssessmentResponse.objects.filter(
                user__in=employee_queryset.values_list("user", flat=True)
            )
            .values("user")
            .distinct()
            .count()
        )

        # Count unique users who have tracked their mood
        mood_tracking_count = (
            MoodTracking.objects.filter(
                user__in=employee_queryset.values_list("user", flat=True)
            )
            .values("user")
            .distinct()
            .count()
        )

        # Count unique users who have used the AI Chatbot (Sana)
        ai_chatbot_count = 0
        try:
            # from sana_ai.models import ChatSession  # TODO: Add sana_ai app
            pass  # Placeholder until sana_ai app is added

            # ai_chatbot_count = (
            #     ChatSession.objects.filter(
            #         user__in=employee_queryset.values_list("user", flat=True)
            #     )
            #     .values("user")
            #     .distinct()
            #     .count()
            # )
        except:
            pass

        # Count unique users who have saved resources
        resource_usage_count = (
            SavedResource.objects.filter(
                user__in=employee_queryset.values_list("user", flat=True)
            )
            .values("user")
            .distinct()
            .count()
        )

        # Calculate percentage of active employees using each feature
        feature_usage = {
            "wellness_assessments": int(
                (wellness_assessments_count / total_users) * 100
            ),
            "ai_chatbot": int((ai_chatbot_count / total_users) * 100),
            "mood_tracking": int((mood_tracking_count / total_users) * 100),
            "resource_library": int((resource_usage_count / total_users) * 100),
        }

        # ============================================================
        # METRIC 7: MOOD TREND (12-WEEK HISTORY)
        # ============================================================
        # Purpose: Shows employee engagement with mood tracking over time
        #
        # How it works:
        # - Tracks number of mood check-ins per week for last 12 weeks
        # - Higher numbers indicate better engagement
        #
        # Why it matters for employers:
        # - Identify engagement patterns (e.g., drops during busy seasons)
        # - Measure effectiveness of wellness initiatives
        # - Spot trends that may need attention
        # - Understand when employees are most engaged with mental health tools
        mood_trend = []
        for week in range(12, 0, -1):
            week_start = datetime.now() - timedelta(weeks=week)
            week_end = week_start + timedelta(weeks=1)

            # Count mood tracking entries for this week
            mood_count = MoodTracking.objects.filter(
                user__in=employee_queryset.values_list("user", flat=True),
                checked_in_at__gte=week_start,
                checked_in_at__lt=week_end,
            ).count()

            mood_trend.append({"week": 13 - week, "value": mood_count})

        # ============================================================
        # METRIC 8: RECENT NOTIFICATIONS
        # ============================================================
        # Purpose: Shows recent organizational activities and updates
        #
        # Examples of notifications:
        # - New employee onboarding
        # - Department changes
        # - System updates
        # - Important announcements
        #
        # Why it matters for employers:
        # - Stay informed about organizational changes
        # - Track important events
        # - Monitor system activity
        # - Quick access to recent updates
        notifications = (
            OrganizationActivity.objects.filter(employer=employer).order_by(
                "-created_at"
            )[:5]
            if employer
            else []
        )

        notifications_data = [
            {
                "id": notif.id,
                "message": notif.description,
                "department": notif.department.name if notif.department else "General",
                "timestamp": notif.created_at,
                "time_ago": self._get_time_ago(notif.created_at),
            }
            for notif in notifications
        ]

        # ============================================================
        # COMPILE ALL DASHBOARD DATA
        # ============================================================
        # This data structure provides everything an employer needs
        # to understand their organization's mental health platform usage
        # and make informed decisions about employee wellbeing initiatives
        data = {
            "summary": {
                "total_employees": total_employees,
                "wellness_index": wellness_index,
                "wellness_description": wellness_description,
                "active_employees": active_employees_count,
            },
            "employees": {
                "list": employees_data,
                "total": total_employees,
            },
            "engagement_trend": {
                "active": engagement_stats["active"],
                "inactive": engagement_stats["inactive"],
                "total": total_employees,
            },
            "feature_usage": feature_usage,
            "mood_trend": mood_trend,
            "notifications": notifications_data,
        }

        return Response(data)

    def _get_time_ago(self, timestamp):
        """
        Helper method to calculate human-readable time differences

        Purpose: Converts timestamps to user-friendly format
        Examples: "2 hours ago", "3 days ago", "just now"
        """
        now = datetime.now()
        if timestamp.tzinfo:
            from django.utils import timezone

            now = timezone.now()

        diff = now - timestamp

        if diff.days > 0:
            return f"{diff.days} day{'s' if diff.days > 1 else ''} ago"
        elif diff.seconds >= 3600:
            hours = diff.seconds // 3600
            return f"{hours} hour{'s' if hours > 1 else ''} ago"
        elif diff.seconds >= 60:
            minutes = diff.seconds // 60
            return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
        else:
            return "just now"


@extend_schema(tags=["Employer Dashboard"])
class EmployeeManagementView(viewsets.ModelViewSet):
    """
    EMPLOYER DASHBOARD - EMPLOYEE MANAGEMENT

    This view handles all employee management operations for employers.

    Purpose for Employers:
    - View complete list of all employees
    - Search employees by name, email, or department
    - Filter employees by department or status
    - Add new employees to the organization
    - Update employee information
    - Remove employees from the system

    Features:
    - Full CRUD operations (Create, Read, Update, Delete)
    - Advanced search functionality
    - Department-based filtering
    - Status-based filtering (active, inactive, suspended)
    - Sorting by multiple fields

    Use Cases:
    - Onboarding new employees
    - Updating employee departments
    - Managing employee status
    - Finding specific employees quickly
    - Generating employee reports
    """

    queryset = Employee.objects.select_related("employer", "department").all()
    serializer_class = EmployeeManagementSerializer
    permission_classes = [IsCompanyAdmin]

    # Search Configuration
    # Employers can search by: first name, last name, email, or department name
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["first_name", "last_name", "email", "department__name"]

    # Sorting Configuration
    # Employers can sort by: name, email, join date, or status
    ordering_fields = ["first_name", "email", "joined_date", "status"]
    ordering = ["-joined_date"]  # Default: newest employees first

    def get_queryset(self):
        """
        Filter employees based on query parameters

        Available Filters:
        - department: Filter by department name (partial match)
        - status: Filter by employee status (exact match)

        Examples:
        - /api/employees/?department=Engineering
        - /api/employees/?status=active
        - /api/employees/?department=HR&status=active
        """
        queryset = super().get_queryset()
        department = self.request.query_params.get("department", None)
        status = self.request.query_params.get("status", None)

        if department:
            queryset = queryset.filter(department__name__icontains=department)
        if status:
            queryset = queryset.filter(status=status)

        return queryset

    def perform_create(self, serializer):
        """
        Create a new employee record

        Purpose: Add new employees to the organization
        Note: Employee will be automatically assigned to the employer's organization
        """
        serializer.save()


@extend_schema(tags=["Employer Dashboard"])
class DepartmentManagementView(viewsets.ModelViewSet):
    """
    EMPLOYER DASHBOARD - DEPARTMENT MANAGEMENT

    This view handles organizational department structure management.

    Purpose for Employers:
    - Create and manage organizational departments
    - Organize employees by department
    - Track department-level metrics
    - Structure the organization hierarchy

    Features:
    - Create new departments
    - Update department information
    - Delete departments
    - Search departments by name
    - Sort departments alphabetically or by creation date

    Use Cases:
    - Setting up organizational structure
    - Reorganizing company departments
    - Tracking department growth
    - Assigning employees to departments

    Examples of Departments:
    - Engineering, HR, Marketing, Sales, Operations, Finance, etc.
    """

    queryset = Department.objects.select_related("employer").all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsCompanyAdmin]

    # Search Configuration
    # Employers can search departments by name
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name"]

    # Sorting Configuration
    # Sort by name (alphabetically) or creation date
    ordering_fields = ["name", "created_at"]
    ordering = ["name"]  # Default: alphabetical order

    def perform_create(self, serializer):
        """
        Create a new department

        Purpose: Add new departments to organize the workforce
        Note: Department is automatically linked to the employer's organization
        """
        # Get employer from request user or use first employer
        employer = Employer.objects.first()
        if hasattr(self.request.user, "employer_profile"):
            employer = self.request.user.employer_profile.employer
        serializer.save(employer=employer)


@extend_schema(tags=["Employer Dashboard"])
class SubscriptionManagementView(viewsets.ModelViewSet):
    """
    EMPLOYER DASHBOARD - SUBSCRIPTION MANAGEMENT

    This view handles all subscription and billing operations for employers.

    Purpose for Employers:
    - View current subscription plan details
    - Browse available subscription plans
    - Upgrade or downgrade subscription
    - View billing history
    - Manage payment methods
    - Track subscription usage (seats used vs available)

    Features:
    - Current subscription overview
    - Plan comparison
    - Billing history with invoices
    - Payment method management
    - Seat usage tracking
    - Renewal date monitoring

    Use Cases:
    - Checking current plan limits
    - Upgrading when adding more employees
    - Reviewing past invoices
    - Updating payment information
    - Planning budget for renewals

    Subscription Plans Typically Include:
    - Starter Plan: Small teams (5-50 employees)
    - Enterprise Plan: Medium to large organizations (50-500 employees)
    - Enterprise Plus: Large corporations (500+ employees)
    """

    queryset = Subscription.objects.select_related(
        "employer", "plan_details", "payment_method"
    ).all()
    serializer_class = SubscriptionManagementSerializer
    """permission_classes = [IsSystemAdminororga]"""

    # Sorting Configuration
    # Sort by start date or amount
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["start_date", "amount"]
    ordering = ["-start_date"]  # Default: newest subscriptions first

    @action(detail=False, methods=["get"])
    def current_subscription(self, request):
        """
        Get Current Active Subscription

        Purpose: View details of the organization's current subscription

        Returns:
        - Plan name and type
        - Number of seats (total and used)
        - Subscription amount
        - Start and end dates
        - Renewal date
        - Payment method
        - Active status

        Why it matters for employers:
        - Know how many employee seats are available
        - Track when renewal is due
        - Understand current plan limitations
        - Plan for upgrades if needed
        """
        subscription = self.get_queryset().filter(is_active=True).first()
        if subscription:
            return Response(SubscriptionManagementSerializer(subscription).data)
        return Response(
            {"message": "No active subscription found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    @action(detail=False, methods=["get"])
    def available_plans(self, request):
        """
        Get Available Subscription Plans

        Purpose: Browse all available subscription plans for comparison

        Returns:
        - Plan names and descriptions
        - Pricing for each plan
        - Number of seats included
        - Features included in each plan

        Why it matters for employers:
        - Compare plans before upgrading
        - Understand what features are available
        - Make informed decisions about plan changes
        - Budget for future growth
        """
        plans = SubscriptionPlan.objects.filter(is_active=True)
        return Response(SubscriptionPlanSerializer(plans, many=True).data)

    @action(detail=False, methods=["get"])
    def billing_history(self, request):
        """
        Get Billing History

        Purpose: View past invoices and payment records

        Returns:
        - Invoice numbers
        - Payment amounts
        - Billing dates
        - Payment status (paid, pending, failed)
        - Plan names for each billing period

        Why it matters for employers:
        - Track expenses for accounting
        - Verify payments
        - Download invoices for records
        - Monitor payment patterns
        - Budget planning
        """
        billing_history = BillingHistory.objects.select_related("employer").order_by(
            "-billing_date"
        )[:10]
        return Response(BillingHistorySerializer(billing_history, many=True).data)


# FLUTTERWAVE WEBHOOK
@api_view(["POST"])
@csrf_exempt
@permission_classes([AllowAny])
def flutterwave_webhook_listener(request):
    """
    Receives and processes webhook events from Flutterwave, verifying the signature.
    """
    # 1. Get raw request body (CRITICAL for HMAC verification)
    raw_body = request.body

    # 2. Get the signature sent by Flutterwave
    flw_signature = request.headers.get("VERIF-HASH")

    # --- SECURITY CHECK ---

    if not flw_signature:
        # No signature header present: discard immediately.
        return HttpResponse(status=401)

    try:
        local_secret = settings.FLW_WEBHOOK_HASH.encode("utf-8")

        # 1. Calculate the raw hash digest (bytes)
        computed_hash_bytes = hmac.new(
            local_secret, raw_body, digestmod=hashlib.sha256
        ).digest()

        # 2. CRITICAL FIX: Encode the bytes to a Base64 string to match Flutterwave's header
        computed_signature = base64.b64encode(computed_hash_bytes).decode("utf-8")

        # 3. Compare the computed Base64 string with the one from Flutterwave
        if computed_signature != flw_signature:
            print("WEBHOOK ERROR: Signature mismatch.")
            return HttpResponse(status=401)  # Unauthorized (Hash verification failed)

    except Exception as e:
        print(f"WEBHOOK HASH ERROR: {e}")
        # Return 500 for a server error during hash computation, or 401 if you treat it as unauthorized
        return HttpResponse(status=401)

    # --- PROCESSING STARTS AFTER SUCCESSFUL VERIFICATION ---
    # ... (Rest of your code remains the same)

    try:
        # ... your transaction verification logic ...
        return HttpResponse(status=200)

    except Exception as e:
        print(f"WEBHOOK PROCESSING ERROR: {e}")
        return HttpResponse(status=500)


#  FOR FRONTEND REDIRECT VERIFICATION
# NOTE: This endpoint is called by your React component after the user is redirected
# back from Flutterwave's payment page. It requires authentication.
@api_view(["POST"])
@permission_classes([AllowAny])
def verify_payment_and_activate_subscription(request):
    """Endpoint called by the frontend (POST) to verify payment and activate subscription."""
    serializer = PaymentVerificationSerializer(data=request.data)
    tx_ref = request.data.get("tx_ref")
    subscription_id = request.data.get(
        "subscription_id"
    )  # Use the ID created in initiate_subscription_payment

    if not tx_ref or not subscription_id:
        return Response(
            {"detail": "Missing transaction reference or subscription ID."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:

        expected_amount = 99.00
        currency = "USD"

        # 2. Verify with Flutterwave
        is_verified, fw_data = verify_flutterwave_transaction(
            tx_ref, expected_amount, currency
        )

        if is_verified:
            # 3. Fulfill the subscription
            with transaction.atomic():
                # sub.is_active = True; sub.save()
                # Invoice.objects.create(...)

                return Response(
                    {"message": "Payment successful and subscription activated."},
                    status=status.HTTP_200_OK,
                )
        else:
            return Response(
                {"detail": fw_data.get("detail", "Verification failed.")},
                status=status.HTTP_400_BAD_REQUEST,
            )

    except Exception as e:
        return Response(
            {"detail": f"Server error during verification: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


#
@action(detail=False, methods=["post"], permission_classes=[IsAuthenticated])
def initiate_subscription_payment(self, request):
    """Initiate payment for a new subscription plan."""

    # --- 1. Validate Input Data and Look up Plan ---
    serializer = SubscriptionInitiateSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    plan_id = serializer.validated_data["plan_id"]
    selected_plan = serializer.selected_plan  # Fetched by the serializer

    try:
        # --- 2. Get User/Organization context (CRITICAL CHANGE HERE) ---
        # The user's 'organization' field already holds the related object.
        organization = request.user.organization

        if not organization:

            return Response(
                {
                    "organization": "User must be linked to an organization to initiate payment."
                },
                status=status.HTTP_403_FORBIDDEN,  # 403 is
            )

        expected_amount = selected_plan.price
        currency = selected_plan.currency

        # Determine seats and initial dates (Simplified calculation)
        seats = (
            selected_plan.seats
        )  # Using 'seats' field from your SubscriptionPlan model
        start_date = timezone.now().date()
        end_date = start_date + timedelta(days=30)
        renewal_date = end_date

        with transaction.atomic():

            # Generate the unique reference ID for the payment gateway
            pending_sub_id = f"SUB-{uuid.uuid4().hex[:8]}"

            # Create the PENDING Subscription record in the database
            pending_sub = Subscription.objects.create(
                employer=organization,
                plan=plan_id,
                plan_details=selected_plan,
                amount=expected_amount,
                seats=seats,
                start_date=start_date,
                end_date=end_date,
                renewal_date=renewal_date,
                is_active=False,
                subscription_reference=pending_sub_id,
            )

            # 4. Call Flutterwave to get the payment link
            payment_result = initiate_payment_fw(
                amount=expected_amount,
                email=request.user.email,
                subscription_id=pending_sub.subscription_reference,
                currency=currency,
            )

        if payment_result.get("status") == "success":
            return Response(
                {
                    "status": "success",
                    "payment_url": payment_result["data"]["link"],
                    "subscription_id": pending_sub.subscription_reference,
                },
                status=status.HTTP_200_OK,
            )
        else:
            return Response(payment_result, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        return Response(
            {"error": f"An unexpected error occurred: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# Wellness Reports View
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
import io, csv

# Import your models and serializers
from .models import CommonIssue, ResourceEngagement, MoodTracking, Department, WellnessGraph
from .serializers import (
    ChatEngagementSerializer,
    DepartmentContributionSerializer,
    OrganizationActivitySerializer,
    WellnessGraphSerializer,
)


@extend_schema(tags=["Employer Dashboard"])
class WellnessReportsView(viewsets.ViewSet):
    """Wellness reports and analytics"""

    permission_classes = [IsAuthenticated]

    # JSON summary endpoint
    def list(self, request):
        """Return wellness summary as JSON"""

        # Metrics
        common_issues = CommonIssue.objects.count()
        resource_engagement = ResourceEngagement.objects.filter(completed=True).count()

        # Average wellbeing using mood scores
        MOOD_SCORES = {
            "Ecstatic": 5,
            "Happy": 4,
            "Excited": 4,
            "Content": 3,
            "Calm": 3,
            "Neutral": 3,
            "Tired": 2,
            "Anxious": 1,
            "Stressed": 1,
            "Sad": 1,
            "Frustrated": 1,
            "Angry": 0,
        }
        moods = MoodTracking.objects.values_list("mood", flat=True)
        if moods:
            avg_wellbeing = round(sum(MOOD_SCORES[m] for m in moods) / len(moods), 2)
        else:
            avg_wellbeing = 0

        at_risk = Department.objects.filter(at_risk=True).count()

        # JSON response
        data = {
            "common_issues": common_issues,
            "resource_engagement": resource_engagement,
            "average_wellbeing_trend": f"{avg_wellbeing:.2f}",
            "at_risk": at_risk,
        }

        return Response(data)

    # CSV download endpoint

    @action(detail=False, methods=["get"], url_path="download-summary")
    def download_summary(self, request):
        """
        Download wellness summary as a professional XLSX file.
        Columns auto-sized for readability.
        """

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Wellness Summary"

        # Add report title
        ws.merge_cells("A1:B1")
        ws["A1"] = "Wellness Summary Report"

        # Leave a blank row for spacing
        ws.append([])

        # Add header
        ws.append(["Metric", "Value"])

        # Compute metrics
        common_issues = CommonIssue.objects.count()
        resource_engagement = ResourceEngagement.objects.filter(completed=True).count()
        MOOD_SCORES = {
            "Ecstatic": 5,
            "Happy": 4,
            "Excited": 4,
            "Content": 3,
            "Calm": 3,
            "Neutral": 3,
            "Tired": 2,
            "Anxious": 1,
            "Stressed": 1,
            "Sad": 1,
            "Frustrated": 1,
            "Angry": 0,
        }
        moods = MoodTracking.objects.values_list("mood", flat=True)
        avg_wellbeing = (
            round(sum(MOOD_SCORES[m] for m in moods) / len(moods), 2) if moods else 0
        )
        at_risk = Department.objects.filter(at_risk=True).count()

        # Write metrics to sheet
        ws.append(["Common Issues", common_issues])
        ws.append(["Resource Engagement", resource_engagement])
        ws.append(["Average Wellbeing Trend", f"{avg_wellbeing:.2f}"])
        ws.append(["At-Risk Departments", at_risk])

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

        # Return XLSX response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="wellness_summary.xlsx"'
        wb.save(response)
        return response


@extend_schema(tags=["Employer Dashboard"])
class OrganizationSettingsView(viewsets.ModelViewSet):
    """Organization settings management"""

    queryset = OrganizationSettings.objects.select_related("employer").all()
    serializer_class = OrganizationSettingsSerializer
    permission_classes = [IsCompanyAdmin]

    def get_queryset(self):
        return OrganizationSettings.objects.filter(employer__user=self.request.user)

    def perform_create(self, serializer):
        # Get or create employer for the current user
        employer, created = Employer.objects.get_or_create(
            name=f"{self.request.user.username}'s Organization"
        )
        serializer.save(employer=employer)


@extend_schema(tags=["Employer Dashboard"])
class TestsByTypeView(viewsets.ViewSet):
    """Tests by type analytics"""

    permission_classes = [IsCompanyAdmin]

    def list(self, request):
        tests_by_type = (
            MoodTracking.objects.values("test_type")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        return Response(list(tests_by_type))


@extend_schema(tags=["Employer Dashboard"])
class TestsByDepartmentView(viewsets.ViewSet):
    """Tests by department analytics"""

    permission_classes = [IsCompanyAdmin]

    def list(self, request):
        tests_by_department = (
            MoodTracking.objects.values("department__name")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        return Response(list(tests_by_department))


# System Admin Views
@extend_schema(tags=["System Admin"])
class SystemAdminOverviewView(viewsets.ViewSet):
    """System Admin dashboard overview"""

    permission_classes = [IsCompanyAdmin]
    serializer_class = SystemAdminOverviewSerializer

    def list(self, request=None):
        from datetime import datetime, timedelta
        from django.db.models import Count

        # Calculate real-time metrics from database
        total_organizations = Employer.objects.count()
        total_clients = Employee.objects.count()

        # Calculate monthly revenue from active subscriptions
        monthly_revenue = (
            Subscription.objects.filter(is_active=True).aggregate(total=Sum("amount"))[
                "total"
            ]
            or 0.00
        )

        # Get hotline calls today
        today = timezone.now().date()
        hotline_calls_today = HotlineCall.objects.filter(call_date__date=today).count()

        # Get organizations this month
        first_day_of_month = datetime.now().replace(day=1).date()
        organizations_this_month = Employer.objects.filter(
            joined_date__gte=first_day_of_month
        ).count()

        # Get clients this month
        clients_this_month = Employee.objects.filter(
            joined_date__gte=first_day_of_month
        ).count()

        # Calculate revenue growth (compare to last month)
        last_month = (datetime.now().replace(day=1) - timedelta(days=1)).replace(day=1)
        last_month_revenue = (
            BillingHistory.objects.filter(
                billing_date__year=last_month.year,
                billing_date__month=last_month.month,
                status="paid",
            ).aggregate(total=Sum("amount"))["total"]
            or 1
        )

        current_month_revenue = (
            BillingHistory.objects.filter(
                billing_date__year=datetime.now().year,
                billing_date__month=datetime.now().month,
                status="paid",
            ).aggregate(total=Sum("amount"))["total"]
            or 0
        )

        revenue_growth_percentage = (
            ((current_month_revenue - last_month_revenue) / last_month_revenue * 100)
            if last_month_revenue > 0
            else 0
        )

        # Calculate hotline growth
        yesterday = today - timedelta(days=1)
        yesterday_calls = HotlineCall.objects.filter(call_date__date=yesterday).count()
        hotline_growth_percentage = (
            ((hotline_calls_today - yesterday_calls) / yesterday_calls * 100)
            if yesterday_calls > 0
            else 0
        )

        # Get platform usage data
        platform_usage = PlatformUsage.objects.all().order_by("week_number")[:6]

        # Get subscription revenue data
        subscription_revenue = SubscriptionRevenue.objects.all().order_by(
            "year", "month"
        )[:9]

        # Get recent system activities
        recent_activities = SystemActivity.objects.select_related(
            "organization"
        ).order_by("-created_at")[:5]

        data = {
            "total_organizations": total_organizations,
            "total_clients": total_clients,
            "monthly_revenue": float(monthly_revenue),
            "hotline_calls_today": hotline_calls_today,
            "organizations_this_month": organizations_this_month,
            "clients_this_month": clients_this_month,
            "revenue_growth_percentage": round(float(revenue_growth_percentage), 2),
            "hotline_growth_percentage": round(float(hotline_growth_percentage), 2),
            "platform_usage": PlatformUsageSerializer(platform_usage, many=True).data,
            "subscription_revenue": SubscriptionRevenueSerializer(
                subscription_revenue, many=True
            ).data,
            "recent_activities": SystemActivitySerializer(
                recent_activities, many=True
            ).data,
        }

        return Response(data)


@extend_schema(tags=["System Admin"])
class OrganizationsManagementView(viewsets.ModelViewSet):
    """Organizations management for system admin"""

    queryset = Employer.objects.all()
    serializer_class = OrganizationsManagementSerializer
    permission_classes = [IsCompanyAdmin]
    filter_backends = [
        filters.SearchFilter,
        filters.OrderingFilter,
        DjangoFilterBackend,
    ]
    search_fields = ["name"]
    ordering_fields = ["name", "joined_date"]
    ordering = ["-joined_date"]
    filterset_fields = ["is_active", "name"]

    def retrieve(self, request, *args, **kwargs):
        """Get detailed information about a specific organization"""
        organization = self.get_object()

        # Get comprehensive organization data
        data = {
            "id": organization.id,
            "name": organization.name,
            "is_active": organization.is_active,
            "joined_date": organization.joined_date,
            # Employee statistics
            "total_employees": organization.employees.count(),
            "active_employees": organization.employees.filter(status="active").count(),
            "inactive_employees": organization.employees.filter(
                status="inactive"
            ).count(),
            # Department information
            "total_departments": organization.departments.count(),
            "at_risk_departments": organization.departments.filter(
                at_risk=True
            ).count(),
            "departments": DepartmentSerializer(
                organization.departments.all(), many=True
            ).data,
            # Subscription information
            "active_subscriptions": organization.subscriptions.filter(
                is_active=True
            ).count(),
            "current_subscription": (
                SubscriptionSerializer(
                    organization.subscriptions.filter(is_active=True).first()
                ).data
                if organization.subscriptions.filter(is_active=True).exists()
                else None
            ),
            # Activity metrics
            "total_assessments": Assessment.objects.filter(
                employee__employer=organization
            ).count(),
            "recent_activities": RecentActivitySerializer(
                organization.activities.order_by("-timestamp")[:10], many=True
            ).data,
            # Engagement metrics
            "engagement_rate": (
                organization.engagements.order_by("-month").first().engagement_rate
                if organization.engagements.exists()
                else 0
            ),
            # Hotline activity
            "hotline_calls": organization.hotline_activities.count(),
            "latest_hotline_activity": (
                HotlineActivitySerializer(
                    organization.hotline_activities.order_by("-recorded_at").first()
                ).data
                if organization.hotline_activities.exists()
                else None
            ),
            # Billing information
            "total_revenue": organization.billing_history.filter(
                status="paid"
            ).aggregate(total=Sum("amount"))["total"]
            or 0,
            "pending_invoices": organization.billing_history.filter(
                status="pending"
            ).count(),
        }

        return Response(data)

    @action(detail=False, methods=["get"], url_path="search-by-name")
    def search_by_name(self, request):
        """Search organizations by name and return detailed results"""
        name = request.query_params.get("name", "")

        if not name:
            return Response(
                {"error": "Please provide a 'name' query parameter"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Search for organizations matching the name
        organizations = Employer.objects.filter(name__icontains=name)

        if not organizations.exists():
            return Response(
                {"message": f"No organizations found matching '{name}'"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Return detailed info for each match
        results = []
        for org in organizations:
            results.append(
                {
                    "id": org.id,
                    "name": org.name,
                    "is_active": org.is_active,
                    "joined_date": org.joined_date,
                    "total_employees": org.employees.count(),
                    "active_employees": org.employees.filter(status="active").count(),
                    "total_departments": org.departments.count(),
                    "current_plan": (
                        org.subscriptions.filter(is_active=True).first().plan
                        if org.subscriptions.filter(is_active=True).exists()
                        else "No active plan"
                    ),
                    "detail_url": f"/admin/organizations/{org.id}/",
                }
            )

        return Response({"count": len(results), "results": results})

    @action(detail=False, methods=["get"])
    def growth_chart(self, request):
        """Get organization growth chart data"""
        from django.db.models import Count
        from datetime import datetime, timedelta

        # Get last 9 months of data
        months = []
        for i in range(9):
            date = datetime.now() - timedelta(days=30 * i)
            months.append(date.strftime("%b"))

        # Get organization counts by month
        growth_data = []
        for i in range(9):
            date = datetime.now() - timedelta(days=30 * i)
            count = Employer.objects.filter(joined_date__lte=date).count()
            growth_data.append(count)

        return Response(
            {"months": list(reversed(months)), "counts": list(reversed(growth_data))}
        )

    @action(detail=False, methods=["get"])
    def client_distribution(self, request):
        """Get client distribution by organization"""
        from django.db.models import Count

        distribution = (
            Employer.objects.annotate(client_count=Count("employees"))
            .values("name", "client_count")
            .order_by("-client_count")[:6]
        )

        return Response(list(distribution))


# HOTLINE VIEWS
@extend_schema(tags=["System Admin"])
class HotlineActivityView(viewsets.ViewSet):
    """Hotline activity management for system admin"""

    permission_classes = [IsCompanyAdmin]
    serializer_class = HotlineActivitySerializer

    def list(self, request):
        from django.db.models import Count
        from datetime import datetime, timedelta

        # Get today's calls
        today_calls = HotlineCall.objects.filter(
            call_date__date=timezone.now().date()
        ).count()

        # Get average duration
        avg_duration = (
            HotlineCall.objects.aggregate(avg_duration=Avg("duration_minutes"))[
                "avg_duration"
            ]
            or 0
        )

        # Get active operators count (unique operators who handled calls today)
        active_operators = (
            HotlineCall.objects.filter(call_date__date=timezone.now().date())
            .values("operator_name")
            .distinct()
            .count()
        )

        # Get hourly volume data for today (real data)
        hourly_volume = []
        for hour in range(24):
            hour_start = timezone.now().replace(
                hour=hour, minute=0, second=0, microsecond=0
            )
            hour_end = hour_start + timedelta(hours=1)
            count = HotlineCall.objects.filter(
                call_date__gte=hour_start, call_date__lt=hour_end
            ).count()
            hourly_volume.append(count)

        # Get call reasons distribution
        call_reasons = (
            HotlineCall.objects.values("reason")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Get recent calls
        recent_calls = HotlineCall.objects.select_related("organization").order_by(
            "-call_date"
        )[:10]

        # Get critical cases
        critical_cases = (
            HotlineCall.objects.filter(urgency="critical")
            .select_related("organization")
            .order_by("-call_date")[:5]
        )

        # Get operator performance (real data from today)
        operator_performance = []
        operators = (
            HotlineCall.objects.filter(call_date__date=timezone.now().date())
            .values("operator_name")
            .distinct()
        )

        for op in operators:
            operator_name = op["operator_name"]
            calls = HotlineCall.objects.filter(
                operator_name=operator_name, call_date__date=timezone.now().date()
            )
            total_calls = calls.count()
            resolved_calls = calls.filter(status="resolved").count()
            resolution_rate = (
                (resolved_calls / total_calls * 100) if total_calls > 0 else 0
            )

            operator_performance.append(
                {
                    "name": operator_name,
                    "calls": total_calls,
                    "resolution_rate": round(resolution_rate, 0),
                }
            )

        # Sort by calls descending
        operator_performance = sorted(
            operator_performance, key=lambda x: x["calls"], reverse=True
        )[:10]

        data = {
            "today_calls": today_calls,
            "average_duration": f"{int(avg_duration//60):02d}:{int(avg_duration%60):02d}",
            "active_operators": active_operators,
            "hourly_volume": hourly_volume,
            "call_reasons": list(call_reasons),
            "recent_calls": HotlineCallSerializer(recent_calls, many=True).data,
            "critical_cases": HotlineCallSerializer(critical_cases, many=True).data,
            "operator_performance": operator_performance,
        }

        return Response(data)


# ActiveHotlineView
@extend_schema(tags=["System Admin"])
class AIManagementView(viewsets.ViewSet):
    """AI Management dashboard for system admin"""

    permission_classes = [IsCompanyAdmin]
    serializer_class = AIManagementSerializer

    def list(self, request):
        # Get total recommendations
        try:
            total_recommendations = (
                AIResource.objects.aggregate(total=Sum("recommended_count"))["total"]
                or 0
            )
        except Exception:
            total_recommendations = 0

        # Get average engagement rate
        try:
            avg_engagement = (
                AIResource.objects.aggregate(avg_rate=Avg("engagement_rate"))[
                    "avg_rate"
                ]
                or 0
            )
        except Exception:
            avg_engagement = 0

        # Get AI accuracy score
        try:
            ai_accuracy = (
                AIResource.objects.aggregate(avg_accuracy=Avg("effectiveness_score"))[
                    "avg_accuracy"
                ]
                or 0
            )
        except Exception:
            ai_accuracy = 0

        # Get effectiveness by resource type
        try:
            effectiveness_by_type = (
                AIResource.objects.values("resource_type")
                .annotate(avg_effectiveness=Avg("effectiveness_score"))
                .order_by("-avg_effectiveness")
            )
        except Exception:
            effectiveness_by_type = []

        # Get weekly recommendations (real data from last 6 weeks)
        from datetime import datetime, timedelta

        weekly_recommendations = []
        for i in range(6):
            week_start = timezone.now() - timedelta(weeks=i + 1)
            week_end = timezone.now() - timedelta(weeks=i)
            week_count = RecommendationLog.objects.filter(
                recommended_on__gte=week_start, recommended_on__lt=week_end
            ).count()
            weekly_recommendations.insert(0, week_count)

        # Get resources
        resources = AIResource.objects.filter(is_active=True).order_by(
            "-effectiveness_score"
        )[:10]

        # Get top anxiety triggers from crisis triggers
        from django.db.models import Count

        top_anxiety_triggers = []
        crisis_triggers = (
            CrisisTrigger.objects.values("detected_phrase")
            .annotate(count=Count("id"))
            .order_by("-count")[:5]
        )

        total_triggers = CrisisTrigger.objects.count()
        for trigger in crisis_triggers:
            percentage = (
                (trigger["count"] / total_triggers * 100) if total_triggers > 0 else 0
            )
            top_anxiety_triggers.append(
                {
                    "trigger": trigger["detected_phrase"],
                    "percentage": round(percentage, 0),
                }
            )

        data = {
            "total_recommendations": total_recommendations,
            "average_engagement_rate": round(avg_engagement, 2),
            "ai_accuracy_score": round(ai_accuracy, 2),
            "effectiveness_by_type": list(effectiveness_by_type),
            "weekly_recommendations": weekly_recommendations,
            "resources": AIResourceSerializer(resources, many=True).data,
            "top_anxiety_triggers": top_anxiety_triggers,
        }

        return Response(data)


@extend_schema(tags=["System Admin"])
class ClientEngagementView(viewsets.ViewSet):
    """Client engagement and rewards dashboard"""

    permission_classes = [IsCompanyAdmin]
    serializer_class = ClientEngagementSerializer

    def list(self, request):
        # Get average daily engagement
        try:
            avg_engagement = (
                ClientEngagement.objects.aggregate(
                    avg_engagement=Avg("engagement_rate")
                )["avg_engagement"]
                or 0
            )
        except Exception:
            avg_engagement = 0

        # Get active reward programs
        try:
            active_rewards = RewardProgram.objects.filter(is_active=True).count()
        except Exception:
            active_rewards = 0

        # Get total points awarded
        try:
            total_points = (
                ClientEngagement.objects.aggregate(total_points=Sum("total_points"))[
                    "total_points"
                ]
                or 0
            )
        except Exception:
            total_points = 0

        # Get weekly engagement data (real data from last 7 days)
        from datetime import datetime, timedelta

        weekly_engagement = []
        for i in range(7):
            day = timezone.now() - timedelta(days=6 - i)
            day_engagement = EngagementTracker.objects.filter(
                employee__user__last_login__date=day.date()
            ).count()
            weekly_engagement.append(day_engagement)

        # Get reward redemptions (real data from last 6 months)
        reward_redemptions = []
        for i in range(6):
            month_date = timezone.now() - timedelta(days=30 * i)
            month_redemptions = (
                RewardProgram.objects.filter(
                    is_active=True,
                    created_at__month=month_date.month,
                    created_at__year=month_date.year,
                ).aggregate(total=Sum("redemption_count"))["total"]
                or 0
            )
            reward_redemptions.insert(0, month_redemptions)

        # Get clients
        clients = ClientEngagement.objects.select_related("organization").order_by(
            "-total_points"
        )[:10]

        # Get top rewards
        top_rewards = RewardProgram.objects.filter(is_active=True).order_by(
            "-redemption_count"
        )[:3]

        # Get engagement trends (real data based on time of day)
        from django.db.models import Q

        total_sessions = ChatSession.objects.count()
        morning_sessions = ChatSession.objects.filter(
            started_at__hour__gte=6, started_at__hour__lt=12
        ).count()
        evening_sessions = ChatSession.objects.filter(
            started_at__hour__gte=18, started_at__hour__lt=24
        ).count()
        weekend_sessions = ChatSession.objects.filter(
            started_at__week_day__in=[1, 7]
        ).count()

        engagement_trends = [
            {
                "trend": "Morning Sessions",
                "percentage": round(
                    (
                        (morning_sessions / total_sessions * 100)
                        if total_sessions > 0
                        else 0
                    ),
                    0,
                ),
            },
            {
                "trend": "Evening Sessions",
                "percentage": round(
                    (
                        (evening_sessions / total_sessions * 100)
                        if total_sessions > 0
                        else 0
                    ),
                    0,
                ),
            },
            {
                "trend": "Weekend Activity",
                "percentage": round(
                    (
                        (weekend_sessions / total_sessions * 100)
                        if total_sessions > 0
                        else 0
                    ),
                    0,
                ),
            },
        ]

        # Get streak statistics (real data)
        streak_7_plus = EngagementStreak.objects.filter(streak_count__gte=7).count()
        streak_14_plus = EngagementStreak.objects.filter(streak_count__gte=14).count()
        streak_30_plus = EngagementStreak.objects.filter(streak_count__gte=30).count()

        streak_stats = [
            {"streak": "7+ Day Streak", "active_users": streak_7_plus},
            {"streak": "14+ Day Streak", "active_users": streak_14_plus},
            {"streak": "30+ Day Streak", "active_users": streak_30_plus},
        ]

        data = {
            "average_daily_engagement": round(avg_engagement, 2),
            "active_reward_programs": active_rewards,
            "total_points_awarded": total_points,
            "weekly_engagement": weekly_engagement,
            "reward_redemptions": reward_redemptions,
            "clients": ClientEngagementSerializer(clients, many=True).data,
            "top_rewards": RewardProgramSerializer(top_rewards, many=True).data,
            "engagement_trends": engagement_trends,
            "streak_statistics": streak_stats,
        }

        return Response(data)


@extend_schema(tags=["System Admin"])
class ReportsAnalyticsView(viewsets.ViewSet):
    """Reports and analytics for system admin"""

    permission_classes = [IsCompanyAdmin]

    def list(self, request):
        from datetime import datetime, timedelta
        from django.db.models import Count

        platform_usage_chart = []
        for i in range(9):
            month_date = timezone.now() - timedelta(days=30 * (8 - i))
            month_users = User.objects.filter(
                last_login__month=month_date.month, last_login__year=month_date.year
            ).count()
            platform_usage_chart.append(month_users)

        total_assessments = MentalHealthAssessment.objects.count()
        health_conditions = []

        # Count anxiety cases (GAD-7)
        anxiety_count = MentalHealthAssessment.objects.filter(
            assessment_type__in=["GAD-7", "BOTH"], gad7_total__gte=10
        ).count()

        # Count depression cases (PHQ-9)
        depression_count = MentalHealthAssessment.objects.filter(
            assessment_type__in=["PHQ-9", "BOTH"], phq9_total__gte=10
        ).count()

        if total_assessments > 0:
            health_conditions = [
                {
                    "condition": "Anxiety",
                    "percentage": round((anxiety_count / total_assessments * 100), 0),
                },
                {
                    "condition": "Depression",
                    "percentage": round(
                        (depression_count / total_assessments * 100), 0
                    ),
                },
                {
                    "condition": "Other",
                    "percentage": round(
                        (
                            100
                            - (anxiety_count / total_assessments * 100)
                            - (depression_count / total_assessments * 100)
                        ),
                        0,
                    ),
                },
            ]
        else:
            health_conditions = [{"condition": "No data available", "percentage": 0}]

        # Get available reports
        available_reports = Report.objects.filter(is_active=True).order_by(
            "-generated_date"
        )[:5]

        # Get custom report options
        custom_report_types = [
            "Platform Usage",
            "Health Conditions",
            "Treatment Outcomes",
            "Organization Performance",
        ]

        date_ranges = [
            "Last 7 Days",
            "Last 30 Days",
            "Last 3 Months",
            "Last 6 Months",
            "Last Year",
        ]

        formats = ["PDF", "Excel", "CSV"]

        data = {
            "platform_usage_chart": platform_usage_chart,
            "health_conditions_distribution": health_conditions,
            "available_reports": ReportSerializer(available_reports, many=True).data,
            "custom_report_types": custom_report_types,
            "date_ranges": date_ranges,
            "formats": formats,
        }

        return Response(data)


# Employee Engagement Summary View
@extend_schema(tags=["Employer Dashboard"])
class EmployeeEngagementSummaryView(APIView):
    permission_classes = [IsCompanyAdmin]

    def get(self, request):
        employees = Employee.objects.filter(employer=request.user.employer)

        total = employees.count()
        active = employees.filter(is_active=True).count()
        inactive = total - active

        active_pct = (active / total * 100) if total else 0
        inactive_pct = (inactive / total * 100) if total else 0

        data = {
            "activeEmployees": active,
            "inactiveEmployees": inactive,
            "totalEmployees": total,
            "activePercentage": round(active_pct, 2),
            "inactivePercentage": round(inactive_pct, 2),
        }

        return Response(data)


@extend_schema(tags=["System Admin"])
class SystemSettingsView(viewsets.ModelViewSet):
    """System settings management"""

    queryset = SystemSettings.objects.all()
    serializer_class = SystemSettingsSerializer
    permission_classes = [IsCompanyAdmin]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["setting_name"]
    ordering_fields = ["setting_name", "updated_at"]
    ordering = ["setting_name"]


"""class FeatureFlagsView(viewsets.ModelViewSet):
    queryset = FeatureFlag.objects.all()
    serializer_class = FeatureFlagSerializer
    permission_classes = [IsCompanyAdmin]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'category', 'is_enabled']
    ordering = ['category', 'name']
    
    @action(detail=False, methods=['get'])
    def by_category(self, request):   
        categories = FeatureFlag.objects.values('category').annotate(
            count=Count('id'),
            enabled_count=Count('id', filter=models.Q(is_enabled=True))
        ).order_by('category')
        
        return Response(list(categories))"""


# views for educational resources
@extend_schema(tags=["Educational_Resources"])
class EducationalResourceViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = EducationalResource.objects.all()
    serializer_class = EducationalResourceSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "description"]
    ordering_fields = ["name", "created_at"]
    
    def retrieve(self, request, *args, **kwargs):
        # Track education feature usage when user views a resource
        if request.user.is_authenticated:
            from .utils.feature import FeatureUsageCalculator
            FeatureUsageCalculator.track_feature(request.user, 'education')
        return super().retrieve(request, *args, **kwargs)


class AdminOrReadOnly(BasePermission):
    """
    Custom permission: Read-only for everyone, write for admins only.
    """

    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:  # GET, HEAD, OPTIONS
            return True
        return request.user and request.user.is_staff  # Only admins can write


class VideoViewSet(viewsets.ModelViewSet):  # Full CRUD support
    queryset = Video.objects.filter(is_active=True)
    serializer_class = VideoSerializer
    permission_classes = [AdminOrReadOnly]  #  Restrict edits/deletes to admins
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["category"]
    search_fields = ["title", "description"]
    ordering_fields = ["created_at", "views", "title"]
    lookup_field = "pk"  # Explicit lookup

    @action(detail=True, methods=["post"])
    def watch(self, request, pk=None):
        """Record that user watched this video"""
        try:
            video = self.get_object()
        except Exception:
            raise NotFound("No video matches the given query.")

        video.views += 1
        video.save()

        if request.user.is_authenticated:
            UserActivity.objects.create(user=request.user, video=video)

        # Notify employees (optional, but should be outside Response)
        for employee in Employee.objects.all():
            Notification.objects.create(
                employee=employee,
                message=f"New video added: {video.title}",
                content_type="video",
                object_id=video.id,
            )

        return Response({"message": "View recorded", "total_views": video.views})

    @action(detail=True, methods=["post"], permission_classes=[IsAdminUser])
    def save(self, request, pk=None):
        """Save video to user's library"""
        try:
            video = self.get_object()
        except Exception:
            raise NotFound("No video matches the given query.")

        saved, created = SavedResource.objects.get_or_create(
            user=request.user, video=video
        )

        if created:
            return Response({"message": "Video saved to your library"})
        else:
            saved.delete()
            return Response({"message": "Video removed from library"})

    @action(detail=False, methods=["get"])
    def popular(self, request):
        """Return top 10 most viewed videos"""
        popular = self.queryset.order_by("-views")[:10]
        serializer = self.get_serializer(popular, many=True)
        return Response(serializer.data)


class AdminOrReadOnly(BasePermission):
    """
    Read-only for everyone, write access only for admins.
    """

    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:  # GET, HEAD, OPTIONS
            return True
        return request.user and request.user.is_staff  # Only admins can write


class AudioViewSet(viewsets.ModelViewSet):  # Full CRUD support
    queryset = Audio.objects.filter(is_active=True)
    serializer_class = AudioSerializer
    permission_classes = [AdminOrReadOnly]  #  Restrict edits/deletes to admins
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["category"]
    search_fields = ["title", "description"]
    ordering_fields = ["created_at", "plays", "title"]
    lookup_field = "pk"

    @action(detail=True, methods=["post"])
    def play(self, request, pk=None):
        """Record that user played this audio"""
        try:
            audio = self.get_object()
        except Exception:
            raise NotFound("No audio matches the given query.")

        audio.plays += 1
        audio.save()

        if request.user.is_authenticated:
            UserActivity.objects.create(user=request.user, audio=audio)

        # Notify employees (should be outside Response)
        for employee in Employee.objects.all():
            Notification.objects.create(
                employee=employee,
                message=f"New audio published1: {audio.title}",
                content_type="audio",
                object_id=audio.id,
            )

        return Response({"message": "Play recorded", "total_plays": audio.plays})

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def save(self, request, pk=None):
        """Save audio to user's library"""
        try:
            audio = self.get_object()
        except Exception:
            raise NotFound("No audio matches the given query.")

        saved, created = SavedResource.objects.get_or_create(
            user=request.user, audio=audio
        )

        if created:
            return Response({"message": "Audio saved to your library"})
        else:
            saved.delete()
            return Response({"message": "Audio removed from library"})


class AdminOrReadOnly(BasePermission):
    """
    Read-only for everyone, write access only for admins.
    """

    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:  # GET, HEAD, OPTIONS
            return True
        return request.user and request.user.is_staff  # Only admins can write


class ArticleViewSet(viewsets.ModelViewSet):  # Full CRUD support
    queryset = Article.objects.all()
    serializer_class = ArticleSerializer
    permission_classes = [AllowAny]

    # Disabling JWT requirements here.
    authentication_classes = []
    parser_classes = [MultiPartParser, FormParser]

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["category"]
    search_fields = ["title", "content", "excerpt"]
    ordering_fields = ["published_date", "views", "reading_time"]

    @action(detail=True, methods=["post"])
    def read(self, request, slug=None):
        """Record that user read this article"""
        try:
            article = self.get_object()
        except Exception:
            raise NotFound("No article matches the given query.")

        article.views += 1
        article.save()

        # Track user activity if authenticated
        if request.user.is_authenticated:
            UserActivity.objects.create(user=request.user, article=article)

        # Notify employees (should be outside Response)
        for employee in Employee.objects.all():
            Notification.objects.create(
                employee=employee,
                message=f"New article published: {article.title}",
                content_type="article",
                object_id=article.id,
            )

        return Response({"message": "Read recorded", "total_views": article.views})

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def save(self, request, slug=None):
        """Save article to user's library"""
        try:
            article = self.get_object()
        except Exception:
            raise NotFound("No article matches the given query.")

        saved, created = SavedResource.objects.get_or_create(
            user=request.user, article=article
        )

        if created:
            return Response({"message": "Article saved to your library"})
        else:
            saved.delete()
            return Response({"message": "Article removed from library"})

    @action(detail=False, methods=["get"])
    def trending(self, request):
        """Return top 10 most viewed articles"""
        trending = self.queryset.order_by("-views")[:10]
        serializer = self.get_serializer(trending, many=True)
        return Response(serializer.data)


from rest_framework.permissions import (
    IsAdminUser,
    IsAuthenticated,
    SAFE_METHODS,
    BasePermission,
)


class AdminOrReadOnly(BasePermission):
    """
    Read-only for everyone, write access only for admins.
    """

    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:  # GET, HEAD, OPTIONS
            return True
        return request.user and request.user.is_staff  # Only admins can write


class MeditationTechniqueViewSet(viewsets.ModelViewSet):  # Full CRUD support
    queryset = MeditationTechnique.objects.filter(is_active=True)
    serializer_class = MeditationTechniqueSerializer
    permission_classes = [AdminOrReadOnly]  # 👈 Restrict edits/deletes to admins
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["category", "difficulty"]
    search_fields = ["title", "description", "benefits"]
    ordering_fields = ["difficulty", "duration", "times_practiced"]

    @action(detail=True, methods=["post"])
    def practice(self, request, pk=None):
        """Record that user practiced this meditation technique"""
        try:
            meditation = self.get_object()
        except Exception:
            raise NotFound("No meditation technique matches the given query.")

        meditation.times_practiced += 1
        meditation.save()

        # Track user activity if authenticated
        if request.user.is_authenticated:
            UserActivity.objects.create(
                user=request.user, meditation=meditation, completed=True
            )

        # Notify employees (should be outside Response)
        for employee in Employee.objects.all():
            Notification.objects.create(
                employee=employee,
                message=f"New meditation published: {meditation.title}",
                content_type="meditation",
                object_id=meditation.id,
            )

        return Response(
            {
                "message": "Practice recorded",
                "total_sessions": meditation.times_practiced,
            }
        )

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated])
    def save(self, request, pk=None):
        """Save meditation technique to user's library"""
        try:
            meditation = self.get_object()
        except Exception:
            raise NotFound("No meditation technique matches the given query.")

        saved, created = SavedResource.objects.get_or_create(
            user=request.user, meditation=meditation
        )

        if created:
            return Response({"message": "Meditation saved to your library"})
        else:
            saved.delete()
            return Response({"message": "Meditation removed from library"})

    @action(detail=False, methods=["get"])
    def for_beginners(self, request):
        """Return beginner-friendly meditation techniques"""
        beginners = self.queryset.filter(difficulty="beginner")
        serializer = self.get_serializer(beginners, many=True)
        return Response(serializer.data)


class SavedResourceViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = SavedResourceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return SavedResource.objects.filter(user=self.request.user)

    @action(detail=False, methods=["get"])
    def by_type(self, request):
        resource_type = request.query_params.get("type", "all")
        saved = self.get_queryset()

        if resource_type == "videos":
            saved = saved.filter(video__isnull=False)
        elif resource_type == "cbt-exercises":
            saved = saved.filter(cbt_exercise__isnull=False)
        elif resource_type == "articles":
            saved = saved.filter(article__isnull=False)
        elif resource_type == "audios":
            saved = saved.filter(audio__isnull=False)
        elif resource_type == "meditations":
            saved = saved.filter(meditation__isnull=False)

        serializer = self.get_serializer(saved, many=True)
        return Response(serializer.data)


class UserActivityViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = UserActivitySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return UserActivity.objects.filter(user=self.request.user)

    @action(detail=False, methods=["get"])
    def stats(self, request):
        activities = self.get_queryset()

        stats = {
            "total_activities": activities.count(),
            "videos_watched": activities.filter(video__isnull=False).count(),
            "cbt_exercises_played": activities.filter(
                cbt_exercise__isnull=False
            ).count(),
            "articles_read": activities.filter(article__isnull=False).count(),
            "audios_listened": activities.filter(audio__isnull=False).count(),
            "meditations_practiced": activities.filter(
                meditation__isnull=False, completed=True
            ).count(),
        }

        return Response(stats)


# @extend_schema(
#     request=OnboardingStateSerializer,
#     responses=OnboardingStateSerializer,
#     methods=["GET", "PATCH"]
# )
# class OnboardingView(RetrieveUpdateAPIView):
#     permission_classes = [IsAuthenticated]
#     serializer_class = OnboardingStateSerializer

#     def get_object(self):
#         obj, _ = OnboardingState.objects.get_or_create(user=self.request.user)
#         return obj

# Note: the onboarding 'state' view was intentionally removed to keep a single
# POST-only completion endpoint. The active `CompleteOnboardingView` is defined
# earlier in this file as a POST-only `APIView` that marks `user.is_onboarded=True`.


@extend_schema_view(
    list=extend_schema(tags=["Dynamic Questions"]),
    retrieve=extend_schema(tags=["Dynamic Questions"]),
    random=extend_schema(
        description="Returns a random set of active dynamic questions.",
        responses=DynamicQuestionSerializer(many=True),
        tags=["Dynamic Questions"],
    ),
)
class DynamicQuestionViewSet(viewsets.ModelViewSet):
    queryset = DynamicQuestion.objects.filter(is_active=True)
    serializer_class = DynamicQuestionSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=["get"])
    def random(self, request):
        count = int(request.query_params.get("count", 5))
        questions = list(self.queryset.order_by("?")[:count])
        serializer = self.get_serializer(questions, many=True)
        return Response(serializer.data)


class UserAchievementViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = UserAchievementSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return UserAchievement.objects.filter(user=self.request.user).select_related(
            "achievement"
        )

    @action(detail=False, methods=["post"])
    def update_progress(self, request):
        """Increment progress for a given achievement"""
        achievement_title = request.data.get("title")
        increment = int(request.data.get("increment", 1))

        achievement = get_object_or_404(
            Achievement, title=achievement_title, is_active=True
        )
        ua, _ = UserAchievement.objects.get_or_create(
            user=request.user, achievement=achievement
        )
        ua.increment_progress(increment)

        serializer = self.get_serializer(ua)
        return Response(serializer.data)
        return UserAchievement.objects.filter(
            user=self.request.user
        ).select_related('achievement')

    @action(detail=False, methods=["get"])
    def summary(self, request):
        """Return overall summary of achievements"""
        achievements = self.get_queryset()
        total = achievements.count()
        completed = achievements.filter(achieved=True).count()
        serializer = self.get_serializer(achievements, many=True)

        return Response(
            {
                "total_achievements": total,
                "completed": completed,
                "progress": serializer.data,
            }
        )

        qs = self.get_queryset()
        total = qs.count()
        completed = qs.filter(achieved=True).count()

        serializer = self.get_serializer(qs, many=True)
        return Response({
            "total_achievements": total,
            "completed": completed,
            "progress": serializer.data
        })

# VIEWS FOR ADMIN USER MANAGEMENT
class OrganizationViewSet(viewsets.ModelViewSet):
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    permission_classes = [IsSystemAdmin]


class AdminUserManagementViewSet(viewsets.ModelViewSet):
    """
    System Admin full control over users.
    """

    queryset = User.objects.all()
    serializer_class = AdminUserSerializer
    permission_classes = [IsSystemAdmin]

    # Temporary suspension
    @action(detail=True, methods=["post"])
    def suspend(self, request, pk=None):
        user = self.get_object()
        if user.role == "system_admin":
            return Response(
                {"detail": "Cannot suspend a system admin."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user.is_suspended = True
        user.save(update_fields=["is_suspended"])
        return Response({"status": "user suspended"})

    # Activate (clear suspension + ensure not deactivated)
    @action(detail=True, methods=["post"])
    def activate(self, request, pk=None):
        user = self.get_object()
        # Ensure employee has avatar before activation
        if user.role == "employee" and not user.avatar:
            return Response(
                {"detail": "Employee must have an avatar before activation."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user.is_suspended = False
        user.is_active = True
        user.save(update_fields=["is_suspended", "is_active"])
        return Response({"status": "user activated"})

    # Deactivate (soft disable) — admin action to remove access
    @action(detail=True, methods=["post"])
    def deactivate(self, request, pk=None):
        user = self.get_object()
        if user.role == "system_admin":
            return Response(
                {"detail": "Cannot deactivate a system admin."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user.is_active = False
        user.is_suspended = False
        user.save(update_fields=["is_active", "is_suspended"])
        return Response({"status": "user deactivated"})

    # Reactivate after deactivation (admin only)
    @action(detail=True, methods=["post"])
    def reactivate(self, request, pk=None):
        user = self.get_object()
        if user.role == "system_admin":
            return Response(
                {"detail": "System admin already active."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if user.role == "employee" and not user.avatar:
            return Response(
                {"detail": "Employee must have an avatar before reactivation."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user.is_active = True
        user.save(update_fields=["is_active"])
        return Response({"status": "user reactivated"})

    # Mark onboarding complete for an employee (admin may set this if needed)
    @action(detail=True, methods=["post"])
    def complete_onboarding(self, request, pk=None):
        user = self.get_object()
        if user.role != "employee":
            return Response(
                {"detail": "Only employees can have onboarding completed."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user.onboarding_completed = True
        user.save(update_fields=["onboarding_completed"])
        return Response({"status": "onboarding completed"})

    # Override destroy to prevent accidental deletion of system_admin accounts
    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        if user.role == "system_admin":
            return Response(
                {"detail": "Cannot delete a system admin account."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)


# ADMIN SUBSCRIPTION MANAGEMENT
class AdminSubscriptionManagementViewSet(viewsets.ModelViewSet):
    queryset = Subscription.objects.all()
    serializer_class = AdminSubscriptionSerializer
    permission_classes = [IsSystemAdmin]


# ADMIN BILLING HISTORY (READ-ONLY)
class AdminBillingHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    """
    System Admin can view billing history but cannot edit.
    """

    queryset = BillingHistory.objects.all()
    serializer_class = AdminBillingSerializer
    permission_classes = [IsSystemAdmin]


# SETTINGS VIEW
@extend_schema(tags=["User - Settings"])
class SettingsViewSet(viewsets.ModelViewSet):
    serializer_class = SettingsSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        # Only return settings for the logged-in user
        return Settings.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        # Ensure settings are tied to the current user
        serializer.save(user=self.request.user)


# media upload view for systems admin
class MediaViewSet(viewsets.ModelViewSet):
    """
    Handles list/retrieve for employees and create/update/delete for system admins.
    """

    queryset = Media.objects.all()
    serializer_class = MediaSerializer
    permission_classes = [IsSystemAdminOrReadOnly]

    # Filtering/searching
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["media_type", "is_published", "tags"]
    search_fields = ["title", "description", "body", "tags"]
    ordering_fields = ["published_at", "created_at"]

    def get_queryset(self):
        qs = super().get_queryset()
        # By default, employees see only published items
        if self.request.method in ("GET",) and not (
            self.request.user.is_staff
            or getattr(self.request.user, "is_system_admin", False)
        ):
            qs = qs.filter(is_published=True)
        return qs

    @action(detail=True, methods=["post"], permission_classes=[IsSystemAdminOrReadOnly])
    def publish(self, request, pk=None):
        """
        Convenience endpoint for system admin to set is_published=True.
        """
        media = self.get_object()
        media.is_published = True
        if not media.published_at:
            from django.utils import timezone

            media.published_at = timezone.now()
        media.save()
        serializer = self.get_serializer(media)
        return Response(serializer.data)

    def perform_destroy(self, instance):
        # Optionally delete files from storage here or rely on storage backend.
        instance.file.delete(save=False)
        instance.thumbnail.delete(save=False)
        instance.delete()


from .permissions import IsAdminOrReadOnly


class JournalEntryViewSet(viewsets.ModelViewSet):
    serializer_class = JournalEntrySerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return JournalEntry.objects.all()
        return JournalEntry.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        # Track journaling feature usage
        from .utils.feature import FeatureUsageCalculator
        FeatureUsageCalculator.track_feature(self.request.user, 'journaling')


from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from .models import CBTExercise
from .serializers import CBTExerciseSerializer
from .permissions import IsAdminOrReadOnly


class CBTExerciseViewSet(viewsets.ModelViewSet):
    serializer_class = CBTExerciseSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]

    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            return CBTExercise.objects.all()
        return CBTExercise.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def perform_update(self, serializer):
        # If marking as completed, auto-set completed_at
        instance = serializer.save()
        if instance.completed and not instance.completed_at:
            from django.utils import timezone

            instance.completed_at = timezone.now()
            instance.save()


from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from drf_spectacular.utils import extend_schema


# --- Payment Method Update ---
class UpdatePaymentMethodViewSet(viewsets.ViewSet):
    """
    API endpoint to receive the new Flutterwave card token and update
    the user's stored payment method for recurring billing.
    """

    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        """POST /payment-methods/"""
        user = request.user
        serializer = PaymentTokenSerializer(data=request.data)

        if serializer.is_valid():
            validated_data = serializer.validated_data
            token_id = validated_data["token_id"]
            card_last_four = validated_data.get("card_last_four")
            card_type = validated_data.get("card_type")
            expiry_month = validated_data.get("expiry_month", 12)
            expiry_year = validated_data.get("expiry_year", 2099)

            try:
                # Get employer from user's employee profile or create a default one
                from obeeomaapp.models import Employer, Employee

                employer = None
                try:
                    employee = Employee.objects.filter(user=user).first()
                    if employee:
                        employer = employee.employer
                except Employee.DoesNotExist:
                    pass

                # If no employer found, we need one for the payment method
                if not employer:
                    return Response(
                        {
                            "detail": "User must be associated with an employer to add payment methods."
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                payment_method, created = PaymentMethod.objects.update_or_create(
                    user=user,
                    defaults={
                        "employer": employer,
                        "token_id": token_id,
                        "last_four_digits": card_last_four or "0000",
                        "card_type": card_type or "Unknown",
                        "expiry_month": expiry_month,
                        "expiry_year": expiry_year,
                        "is_default": True,
                    },
                )

                response_message = (
                    "New payment method successfully created."
                    if created
                    else "Payment method successfully updated."
                )

                return Response(
                    {"message": response_message, "token_id": token_id},
                    status=status.HTTP_200_OK,
                )

            except Exception as e:
                print(f"Database error during payment update: {e}")
                return Response(
                    {
                        "detail": "A server error occurred while saving the payment token."
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["Assessments"])
class PSS10AssessmentViewSet(viewsets.ModelViewSet):
    """
    Full CRUD API endpoint for Perceived Stress Scale (PSS-10).
    Supports: GET, POST, PUT, PATCH, DELETE
    """

    permission_classes = [IsAuthenticated]
    serializer_class = PSS10AssessmentSerializer
    queryset = PSS10Assessment.objects.all()

    def get_queryset(self):
        # Return only assessments for the logged-in user
        return PSS10Assessment.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        responses = serializer.validated_data["responses"]
        score = sum(responses)

        if score <= 13:
            category = "Low stress"
        elif 14 <= score <= 26:
            category = "Moderate stress"
        else:
            category = "High stress"

        serializer.save(user=self.request.user, score=score, category=category)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        instance = serializer.instance
        return Response(
            {
                "score": instance.score,
                "category": instance.category,
                "user": instance.user.id,
                "message": f"Your stress level is {instance.category.lower()}.",
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )


# content/views.py
import uuid
import boto3
from botocore.client import Config
from django.conf import settings
from rest_framework import viewsets, status, permissions, views
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import ContentArticle, ContentMedia, Article
from .serializers import ContentArticleSerializer, ContentMediaSerializer, ArticleSerializer


# --- Helper to create a presigned PUT URL for DO Spaces ---
def generate_presigned_put_url(
    object_key: str, content_type: str, expires_in: int = 3600
):
    client = boto3.client(
        "s3",
        region_name=settings.AWS_S3_REGION_NAME,
        endpoint_url=settings.AWS_S3_ENDPOINT_URL,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        config=Config(signature_version="s3v4"),
    )

    url = client.generate_presigned_url(
        ClientMethod="put_object",
        Params={
            "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
            "Key": object_key,
            "ContentType": content_type,
        },
        ExpiresIn=expires_in,
    )
    return url


# --- Presign upload endpoint (separate view) ---
class PresignUploadView(views.APIView):
    permission_classes = [IsSystemAdmin]

    def post(self, request):
        """
        Expects JSON:
        { "filename": "video.mp4", "content_type": "video/mp4", "media_type": "video", "title": "optional" }
        Returns:
        { "presigned_url": "...", "s3_key": "uploads/uuid_video.mp4", "media_id": 5 }
        """
        filename = request.data.get("filename")
        content_type = request.data.get("content_type")
        media_type = request.data.get("media_type")
        title = request.data.get("title", "")

        if not filename or not content_type:
            return Response(
                {"detail": "filename and content_type are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # create a safe unique object key (you can change path structure)
        unique_name = f"{uuid.uuid4().hex}_{filename}"
        object_key = f"uploads/{unique_name}"

        try:
            presigned_url = generate_presigned_put_url(object_key, content_type)
        except Exception as e:
            return Response(
                {"detail": f"failed to generate presigned url: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # create DB record now (uploaded=False) so we can reference it
        media = ContentMedia.objects.create(
            owner=request.user,
            title=title or filename,
            media_type=media_type or ContentMedia.MEDIA_OTHER,
            s3_key=object_key,
            uploaded=False,
            processed=False,
        )

        return Response(
            {
                "presigned_url": presigned_url,
                "s3_key": object_key,
                
                "media_id": media.id,
            },
            status=status.HTTP_201_CREATED,
        )


# --- Confirm upload endpoint ---
class ConfirmUploadView(views.APIView):
    permission_classes = [IsSystemAdmin]

    def post(self, request):
        """
        Body: { "media_id": 5, "public_url": "optional public url from CDN" }
        Marks the media as uploaded=True and optionally records public_url.
        """
        media_id = request.data.get("media_id")
        public_url = request.data.get("public_url")

        if not media_id:
            return Response(
                {"detail": "media_id required"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            media = ContentMedia.objects.get(id=media_id, owner=request.user)
        except ContentMedia.DoesNotExist:
            return Response(
                {"detail": "media not found"}, status=status.HTTP_404_NOT_FOUND
            )

        media.uploaded = True
        if public_url:
            media.public_url = public_url
        media.save()

        # TODO: enqueue background processing task e.g. tasks.process_media.delay(media.id)
        return Response(
            {"detail": "upload confirmed", "media_id": media.id},
            status=status.HTTP_200_OK,
        )


# --- ViewSets for Article and Media ---
class ContentArticleViewSet(viewsets.ModelViewSet):
    serializer_class = ContentArticleSerializer
    permission_classes = [IsSystemAdminOrReadOnly]

    def get_queryset(self):
        if (
            self.request.user
            and self.request.user.is_authenticated
            and (
                self.request.user.is_superuser
                or self.request.user.role == "system_admin"
            )
        ):
            # System admin sees all articles
            return ContentArticle.objects.all().order_by("-created_at")
        else:
            # Employees see only published articles
            return ContentArticle.objects.filter(published=True).order_by("-created_at")

    def perform_create(self, serializer):
        serializer.save(author=self.request.user)


class ContentMediaViewSet(viewsets.ModelViewSet):
    serializer_class = ContentMediaSerializer
    permission_classes = [IsSystemAdminOrReadOnly]

    def create(self, request, *args, **kwargs):
        """Handle FormData uploads with new fields"""
        # Extract data from FormData
        title = request.data.get('title', '')
        description = request.data.get('description', '')
        category = request.data.get('category', '')
        content_status = request.data.get('status', 'draft')
        duration = request.data.get('duration', '')
        file_size = request.data.get('file_size', '')
        media_type = request.data.get('media_type', 'other')
        
        # Create ContentMedia object with all fields
        content_media = ContentMedia.objects.create(
            owner=request.user,
            title=title,
            description=description,
            category=category if category else None,
            status=content_status,
            duration=duration,
            file_size=file_size,
            media_type=media_type,
            uploaded=True if request.FILES.get('file') else False,
        )
        
        # Handle file upload if present
        if 'file' in request.FILES:
            file = request.FILES['file']
            
            # Create the uploads directory if it doesn't exist
            import os
            uploads_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
            os.makedirs(uploads_dir, exist_ok=True)
            
            # Save the file to disk
            file_path = os.path.join(uploads_dir, file.name)
            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            
            # Update the database record
            content_media.s3_key = f"uploads/{file.name}"
            content_media.uploaded = True
            content_media.save()
        
        serializer = self.get_serializer(content_media)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def get_queryset(self):
        # Allow all access for superuser gideonmuteso@gmail.com
        if (
            self.request.user and 
            self.request.user.is_authenticated and 
            self.request.user.email == "gideonmuteso@gmail.com"
        ):
            # Superuser sees all media without restrictions
            return ContentMedia.objects.all().order_by("-created_at")
        
        # Normal authentication logic for other users
        if (
            self.request.user
            and self.request.user.is_authenticated
            and (
                self.request.user.is_superuser
                or getattr(self.request.user, 'role', None) == "system_admin"
            )
        ):
            # System admin sees all media
            return ContentMedia.objects.all().order_by("-created_at")
        else:
            # Employees see only uploaded media (removed processed filter)
            return ContentMedia.objects.filter(uploaded=True).order_by(
                "-created_at"
            )

    def perform_create(self, serializer):
        # If it's the special superuser, allow creation without auth checks
        if (
            self.request.user and 
            self.request.user.is_authenticated and 
            self.request.user.email == "gideonmuteso@gmail.com"
        ):
            serializer.save(owner=self.request.user)
        else:
            serializer.save(owner=self.request.user)


# New views for the requested endpoints


@extend_schema(tags=["Engagement Level"])
class EngagementLevelViewSet(viewsets.ModelViewSet):
    queryset = EngagementLevel.objects.all()
    serializer_class = EngagementLevelSerializer
    permission_classes = [permissions.IsAuthenticated]


@extend_schema(tags=["Company Mood"])
class CompanyMoodViewSet(viewsets.ModelViewSet):
    serializer_class = CompanyMoodSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return CompanyMood.objects.all()
        
        # Get organization based on user role
        if hasattr(user, 'employer'):
            organization = user.employer.organization
        elif hasattr(user, 'employee'):
            organization = user.employee.organization
        else:
            return CompanyMood.objects.none()
            
        return CompanyMood.objects.filter(organization=organization)

    def perform_create(self, serializer):
        user = self.request.user
        
        # Get organization based on user role
        if hasattr(user, 'employer'):
            organization = user.employer.organization
        elif hasattr(user, 'employee'):
            organization = user.employee.organization
        else:
            raise ValidationError("User must be associated with an organization")
            
        serializer.save(organization=organization)

    @action(detail=False, methods=['post'], url_path='aggregate-today')
    def aggregate_today(self, request):
        """Aggregate today's mood data for the organization"""
        user = request.user
        
        # Get organization
        if hasattr(user, 'employer'):
            organization = user.employer.organization
        elif hasattr(user, 'employee'):
            organization = user.employee.organization
        else:
            return Response(
                {"error": "User must be associated with an organization"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        today = now().date()
        
        # Get all mood entries for today from employees in this organization
        today_entries = MoodTracking.objects.filter(
            employee__organization=organization,
            checked_in_at__date=today
        )

        if not today_entries.exists():
            return Response(
                {"message": "No mood entries found for today"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Calculate aggregated data
        company_mood = self._calculate_company_mood(organization, today, today_entries)
        
        # Create or update CompanyMood record
        mood_record, created = CompanyMood.objects.update_or_create(
            organization=organization,
            date=today,
            defaults=company_mood
        )

        serializer = self.get_serializer(mood_record)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='aggregate-period')
    def aggregate_period(self, request):
        """Aggregate mood data for a specific period"""
        user = request.user
        
        # Get organization
        if hasattr(user, 'employer'):
            organization = user.employer.organization
        elif hasattr(user, 'employee'):
            organization = user.employee.organization
        else:
            return Response(
                {"error": "User must be associated with an organization"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        if not start_date or not end_date:
            return Response(
                {"error": "Both start_date and end_date are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get mood entries for the period
        period_entries = MoodTracking.objects.filter(
            employee__organization=organization,
            checked_in_at__date__range=[start_date, end_date]
        )

        if not period_entries.exists():
            return Response(
                {"message": "No mood entries found for the specified period"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Group entries by date and aggregate
        results = []
        current_date = start_date
        while current_date <= end_date:
            daily_entries = period_entries.filter(checked_in_at__date=current_date)
            
            if daily_entries.exists():
                company_mood = self._calculate_company_mood(organization, current_date, daily_entries)
                
                mood_record, created = CompanyMood.objects.update_or_create(
                    organization=organization,
                    date=current_date,
                    defaults=company_mood
                )
                
                serializer = self.get_serializer(mood_record)
                results.append(serializer.data)
            
            current_date += timedelta(days=1)

        return Response({
            "message": f"Aggregated mood data for {len(results)} days",
            "data": results
        })

    @action(detail=False, methods=['get'], url_path='dashboard-summary')
    def dashboard_summary(self, request):
        """Get comprehensive mood summary for dashboard"""
        user = request.user
        
        # Get organization
        if hasattr(user, 'employer'):
            organization = user.employer.organization
        elif hasattr(user, 'employee'):
            organization = user.employee.organization
        else:
            return Response(
                {"error": "User must be associated with an organization"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get different time periods
        today = now().date()
        last_7_days = today - timedelta(days=7)
        last_30_days = today - timedelta(days=30)

        # Today's data
        today_mood = CompanyMood.objects.filter(
            organization=organization,
            date=today
        ).first()

        # Last 7 days
        recent_moods = CompanyMood.objects.filter(
            organization=organization,
            date__gte=last_7_days
        ).order_by('date')

        # Last 30 days
        monthly_moods = CompanyMood.objects.filter(
            organization=organization,
            date__gte=last_30_days
        ).order_by('date')

        # Calculate trends
        trend_data = self._calculate_trends(recent_moods)

        # Get employee participation
        total_employees = Employee.objects.filter(organization=organization).count()
        today_participants = MoodTracking.objects.filter(
            employee__organization=organization,
            checked_in_at__date=today
        ).values('employee').distinct().count()

        summary = {
            "today": {
                "total_entries": today_mood.total_entries if today_mood else 0,
                "average_mood": today_mood.average_mood_score if today_mood else 0,
                "positive_percentage": today_mood.positive_percentage if today_mood else 0,
                "dominant_mood": today_mood.dominant_mood if today_mood else "No data",
                "participants": today_participants,
                "participation_rate": round((today_participants / total_employees * 100), 1) if total_employees > 0 else 0
            },
            "last_7_days": {
                "total_entries": sum(mood.total_entries for mood in recent_moods),
                "average_mood": sum(mood.average_mood_score for mood in recent_moods) / len(recent_moods) if recent_moods.exists() else 0,
                "positive_percentage": sum(mood.positive_count for mood in recent_moods) / sum(mood.total_entries for mood in recent_moods) * 100 if recent_moods.exists() and sum(mood.total_entries for mood in recent_moods) > 0 else 0,
                "trend": trend_data['trend'],
                "days_with_data": recent_moods.count()
            },
            "last_30_days": {
                "total_entries": sum(mood.total_entries for mood in monthly_moods),
                "average_mood": sum(mood.average_mood_score for mood in monthly_moods) / len(monthly_moods) if monthly_moods.exists() else 0,
                "days_with_data": monthly_moods.count()
            },
            "organization_stats": {
                "total_employees": total_employees,
                "active_days": recent_moods.count(),
                "most_common_mood": self._get_most_common_mood(recent_moods)
            }
        }

        return Response(summary)

    def _calculate_company_mood(self, organization, date, entries):
        """Calculate aggregated mood data for a specific day"""
        # Initialize mood counts
        mood_counts = {
            'Ecstatic': 0, 'Happy': 0, 'Excited': 0, 'Content': 0,
            'Calm': 0, 'Neutral': 0, 'Tired': 0,
            'Anxious': 0, 'Stressed': 0, 'Sad': 0, 'Frustrated': 0, 'Angry': 0
        }

        # Count each mood
        for entry in entries:
            mood = entry.mood
            if mood in mood_counts:
                mood_counts[mood] += 1

        # Calculate category totals
        positive_moods = ['Ecstatic', 'Happy', 'Excited', 'Content']
        neutral_moods = ['Calm', 'Neutral', 'Tired']
        negative_moods = ['Anxious', 'Stressed', 'Sad', 'Frustrated', 'Angry']

        positive_count = sum(mood_counts[mood] for mood in positive_moods)
        neutral_count = sum(mood_counts[mood] for mood in neutral_moods)
        negative_count = sum(mood_counts[mood] for mood in negative_moods)

        total_entries = entries.count()

        # Calculate average mood score (1-5 scale)
        mood_scores = {
            'Ecstatic': 5, 'Happy': 4, 'Excited': 4, 'Content': 3,
            'Calm': 3, 'Neutral': 2, 'Tired': 2,
            'Anxious': 1, 'Stressed': 1, 'Sad': 1, 'Frustrated': 1, 'Angry': 0
        }

        total_score = sum(mood_scores.get(entry.mood, 2) for entry in entries)
        average_score = total_score / total_entries if total_entries > 0 else 0

        # Find dominant mood
        dominant_mood = max(mood_counts, key=mood_counts.get) if total_entries > 0 else 'None'

        # Generate summary description
        summary = self._generate_summary(positive_count, neutral_count, negative_count, dominant_mood, total_entries)

        return {
            'total_entries': total_entries,
            'average_mood_score': round(average_score, 2),
            'ecstatic_count': mood_counts['Ecstatic'],
            'happy_count': mood_counts['Happy'],
            'excited_count': mood_counts['Excited'],
            'content_count': mood_counts['Content'],
            'calm_count': mood_counts['Calm'],
            'neutral_count': mood_counts['Neutral'],
            'tired_count': mood_counts['Tired'],
            'anxious_count': mood_counts['Anxious'],
            'stressed_count': mood_counts['Stressed'],
            'sad_count': mood_counts['Sad'],
            'frustrated_count': mood_counts['Frustrated'],
            'angry_count': mood_counts['Angry'],
            'positive_count': positive_count,
            'neutral_mood_count': neutral_count,
            'negative_count': negative_count,
            'summary_description': summary,
            'dominant_mood': dominant_mood,
            'sentiment_trend': 'stable'  # Will be calculated separately
        }

    def _generate_summary(self, positive, neutral, negative, dominant_mood, total):
        """Generate a human-readable summary of the mood data"""
        if total == 0:
            return "No mood data available for this period."

        positive_pct = (positive / total) * 100
        negative_pct = (negative / total) * 100

        summary_parts = []

        if positive_pct >= 70:
            summary_parts.append("Overall mood is very positive")
        elif positive_pct >= 50:
            summary_parts.append("Overall mood is moderately positive")
        elif negative_pct >= 50:
            summary_parts.append("Overall mood shows concerning negative trends")
        else:
            summary_parts.append("Overall mood is mixed")

        summary_parts.append(f"with {dominant_mood.lower()} being the most common mood.")

        if negative_pct >= 30:
            summary_parts.append("Consider implementing wellness initiatives to support employee mental health.")

        return " ".join(summary_parts)

    def _calculate_trends(self, recent_moods):
        """Calculate mood trends over recent days"""
        if len(recent_moods) < 2:
            return {'trend': 'insufficient_data'}

        # Compare first half vs second half
        mid_point = len(recent_moods) // 2
        first_half = recent_moods[:mid_point]
        second_half = recent_moods[mid_point:]

        first_avg = sum(mood.average_mood_score for mood in first_half) / len(first_half) if first_half else 0
        second_avg = sum(mood.average_mood_score for mood in second_half) / len(second_half) if second_half else 0

        if second_avg > first_avg + 0.3:
            trend = 'improving'
        elif second_avg < first_avg - 0.3:
            trend = 'declining'
        else:
            trend = 'stable'

        return {'trend': trend}

    def _get_most_common_mood(self, moods):
        """Get the most common mood across multiple days"""
        mood_totals = {}
        
        for mood_data in moods:
            mood_totals[mood_data.dominant_mood] = mood_totals.get(mood_data.dominant_mood, 0) + 1

        return max(mood_totals, key=mood_totals.get) if mood_totals else "No data"


@extend_schema(tags=["Wellness Graph"])
class WellnessGraphViewSet(viewsets.ModelViewSet):
    serializer_class = WellnessGraphSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return WellnessGraph.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=['get'], url_path='sync-from-mood-tracking')
    def sync_from_mood_tracking(self, request):
        """Sync wellness graph data from mood tracking entries"""
        user = request.user
        
        # Get all mood tracking entries for this user
        mood_entries = MoodTracking.objects.filter(user=user).order_by('checked_in_at')
        
        if not mood_entries.exists():
            return Response(
                {"message": "No mood tracking entries found to sync"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Convert mood names to numeric scores
        mood_scores = {
            'Ecstatic': 5, 'Happy': 4, 'Excited': 4, 'Content': 3,
            'Calm': 3, 'Neutral': 2, 'Tired': 2,
            'Anxious': 1, 'Stressed': 1, 'Sad': 1, 'Frustrated': 1, 'Angry': 0
        }
        
        synced_count = 0
        updated_count = 0
        
        for entry in mood_entries:
            mood_date = entry.checked_in_at.date()
            mood_score = mood_scores.get(entry.mood, 2)  # Default to neutral (2)
            
            # Create or update wellness graph entry
            wellness_entry, created = WellnessGraph.objects.update_or_create(
                user=user,
                mood_date=mood_date,
                defaults={'mood_score': mood_score}
            )
            
            if created:
                synced_count += 1
            else:
                # Update if score is different
                if wellness_entry.mood_score != mood_score:
                    wellness_entry.mood_score = mood_score
                    wellness_entry.save()
                    updated_count += 1
        
        return Response({
            "message": "Successfully synced mood tracking data to wellness graph",
            "synced_entries": synced_count,
            "updated_entries": updated_count,
            "total_processed": synced_count + updated_count
        })

    @action(detail=False, methods=['get'], url_path='chart-data')
    def chart_data(self, request):
        """Get wellness graph data optimized for frontend charts"""
        user = request.user
        
        # Get date range from query params
        days = int(request.query_params.get('days', 30))
        start_date = now().date() - timedelta(days=days)
        
        # Get wellness graph data
        wellness_data = WellnessGraph.objects.filter(
            user=user,
            mood_date__gte=start_date
        ).order_by('mood_date')
        
        # If no wellness graph data, try to sync from mood tracking
        if not wellness_data.exists():
            sync_response = self.sync_from_mood_tracking(request)
            if sync_response.status_code == 200:
                # Retry getting data after sync
                wellness_data = WellnessGraph.objects.filter(
                    user=user,
                    mood_date__gte=start_date
                ).order_by('mood_date')
        
        # Format data for frontend
        chart_data = []
        for entry in wellness_data:
            chart_data.append({
                'date': entry.mood_date.strftime('%Y-%m-%d'),
                'score': entry.mood_score,
                'mood_label': self._get_mood_label(entry.mood_score)
            })
        
        # Calculate statistics
        if chart_data:
            scores = [item['score'] for item in chart_data]
            avg_score = sum(scores) / len(scores)
            max_score = max(scores)
            min_score = min(scores)
            
            # Determine trend (last 7 days vs previous 7 days)
            if len(chart_data) >= 14:
                recent_avg = sum(scores[-7:]) / 7
                previous_avg = sum(scores[-14:-7]) / 7
                trend = 'improving' if recent_avg > previous_avg else 'declining' if recent_avg < previous_avg else 'stable'
            else:
                trend = 'insufficient_data'
        else:
            avg_score = max_score = min_score = 0
            trend = 'no_data'
        
        return Response({
            "data": chart_data,
            "statistics": {
                "average_score": round(avg_score, 2),
                "max_score": max_score,
                "min_score": min_score,
                "total_days": len(chart_data),
                "trend": trend
            },
            "period": f"last_{days}_days"
        })

    @action(detail=False, methods=['post'], url_path='auto-sync')
    def auto_sync(self, request):
        """Automatically sync latest mood tracking entries"""
        user = request.user
        
        # Get the latest mood tracking entry
        latest_mood = MoodTracking.objects.filter(user=user).order_by('-checked_in_at').first()
        
        if not latest_mood:
            return Response(
                {"message": "No mood tracking entries found"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Convert to wellness graph entry
        mood_scores = {
            'Ecstatic': 5, 'Happy': 4, 'Excited': 4, 'Content': 3,
            'Calm': 3, 'Neutral': 2, 'Tired': 2,
            'Anxious': 1, 'Stressed': 1, 'Sad': 1, 'Frustrated': 1, 'Angry': 0
        }
        
        mood_date = latest_mood.checked_in_at.date()
        mood_score = mood_scores.get(latest_mood.mood, 2)
        
        # Create or update wellness graph entry
        wellness_entry, created = WellnessGraph.objects.update_or_create(
            user=user,
            mood_date=mood_date,
            defaults={'mood_score': mood_score}
        )
        
        return Response({
            "message": "Auto-sync completed",
            "mood_date": mood_date.strftime('%Y-%m-%d'),
            "mood_score": mood_score,
            "original_mood": latest_mood.mood,
            "was_created": created
        })

    def _get_mood_label(self, score):
        """Convert numeric score back to mood label"""
        mood_labels = {
            5: 'Ecstatic',
            4: 'Happy/Excited',
            3: 'Content/Calm',
            2: 'Neutral/Tired',
            1: 'Anxious/Stressed/Sad',
            0: 'Angry/Frustrated'
        }
        return mood_labels.get(score, 'Unknown')


@extend_schema(tags=["Employee Management"])
class EmployeeManagementViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeManagementSerializer
    permission_classes = [permissions.IsAuthenticated]


@extend_schema(tags=["Notifications"])
class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(employee__user=self.request.user)

    # Provide a convenience endpoint to regenerate a presigned GET url for private serving (optional)
    @action(
        detail=True, methods=["get"], permission_classes=[permissions.IsAuthenticated]
    )
    def get_presigned_get(self, request, pk=None):
        media = self.get_object()
        # Only owner or staff can request a signed GET URL
        if media.owner != request.user and not request.user.is_staff:
            return Response({"detail": "forbidden"}, status=status.HTTP_403_FORBIDDEN)

        client = boto3.client(
            "s3",
            region_name=settings.AWS_S3_REGION_NAME,
            endpoint_url=settings.AWS_S3_ENDPOINT_URL,
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            config=Config(signature_version="s3v4"),
        )

        try:
            url = client.generate_presigned_url(
                ClientMethod="get_object",
                Params={
                    "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                    "Key": media.s3_key,
                },
                ExpiresIn=300,  # 5 minutes
            )
        except Exception as e:
            return Response(
                {"detail": f"failed to generate get url: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# Server-side authentication check endpoint
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def auth_check(request):
    """
    Check if user is authenticated on server side
    This is used as a backup for client-side authentication
    """
    return Response({
        'authenticated': True,
        'user': {
            'id': request.user.id,
            'email': request.user.email,
            'role': request.user.role
        }
    })


# Admin AI Chat View
@extend_schema(tags=["Admin - AI Chat"])
class AdminChatView(viewsets.ViewSet):
    """
    Admin-specific AI chat functionality
    No sessions - just message history with last 10 messages for context
    """
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AdminChatMessageSerializer

    def get_queryset(self):
        """Only system admins can access their own chat messages"""
        # Check if user is system admin using User model role field
        if self.request.user.role != "system_admin":
            return AdminChatMessage.objects.none()
        return AdminChatMessage.objects.filter(admin=self.request.user)

    def list(self, request):
        """Get last 10 messages for context"""
        # Check if user is system admin using User model role field
        if request.user.role != "system_admin":
            return Response(
                {"error": "Access denied. System admins only."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        messages = self.get_queryset()[:10]  # Get last 10 messages
        serializer = self.serializer_class(messages, many=True)
        return Response(serializer.data)

    def create(self, request):
        """Send a message and get AI response"""
        # Debug: Check if environment variable is loaded
        logger.info(f"DEBUG: GROQ_API_KEY loaded: {'YES' if os.environ.get('GROQ_API_KEY') else 'MISSING'}")
        
        # Check if user is system admin using User model role field
        if request.user.role != "system_admin":
            return Response(
                {"error": "Access denied. System admins only."},
                status=status.HTTP_403_FORBIDDEN
            )

        user_message = request.data.get("message", "").strip()
        if not user_message:
            return Response(
                {"error": "Message cannot be empty"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Save user message
            admin_message = AdminChatMessage.objects.create(
                admin=request.user,
                sender="admin",
                message=user_message
            )

            # Get last 10 messages for context (excluding the one we just saved)
            conversation_history = []
            recent_messages = self.get_queryset()[:10]
            
            for msg in recent_messages:
                if msg.id != admin_message.id:  # Skip the message we just saved
                    conversation_history.append({
                        "role": msg.api_role,  # Property, not method - no parentheses
                        "content": msg.message
                    })

            # Get AI response using existing GroqService
            logger.info("DEBUG: About to initialize GroqService")
            groq_service = GroqService()
            logger.info("DEBUG: GroqService initialized successfully")
            
            # Enhanced system prompt for admin-specific insights
            system_prompt = """You are an AI assistant for the Obeeoma mental health platform administrator. 
            You provide insights on:
            1. Resource consumption patterns and trends
            2. Platform growth strategies and recommendations
            3. System optimization suggestions
            4. User engagement analytics
            5. Mental health platform best practices
            
            Be concise, data-driven, and actionable in your responses."""
            
            # Add system prompt to conversation history
            full_conversation = [{"role": "system", "content": system_prompt}] + conversation_history
            logger.info(f"DEBUG: About to call Groq API with {len(full_conversation)} messages")
            
            ai_reply = groq_service.get_response(
                user_message=user_message,
                conversation_history=full_conversation,
            )
            logger.info("DEBUG: Groq API call successful")

            # Save AI response
            ai_message = AdminChatMessage.objects.create(
                admin=request.user,
                sender="ai",
                message=ai_reply
            )

            # Return both messages
            response_data = {
                "user_message": AdminChatMessageSerializer(admin_message).data,
                "ai_response": AdminChatMessageSerializer(ai_message).data,
            }

            return Response(response_data, status=status.HTTP_201_CREATED)

        except ValueError as e:
            logger.error(f"Groq API configuration error: {str(e)}")
            return Response(
                {"error": "AI service not configured. Please contact support."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        except Exception as e:
            logger.error(f"Admin AI chat error: {str(e)}")
            return Response(
                {"error": "Failed to process message. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=['delete'])
    def clear_history(self, request):
        """Clear all chat history for this admin"""
        if request.user.role != "system_admin":
            return Response(
                {"error": "Access denied. System admins only."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        AdminChatMessage.objects.filter(admin=request.user).delete()
        return Response(
            {"message": "Chat history cleared successfully"},
            status=status.HTTP_200_OK
        )


# Admin AI Status Management View
@extend_schema(tags=["Admin - AI Status"])
class AdminAIStatusView(viewsets.ViewSet):
    """
    Admin AI status management functionality
    Handle toggling AI features and retrieving status
    """
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AIStatusSerializer

    def list(self, request):
        """Get all AI feature statuses"""
        # Check if user is system admin
        if request.user.role != "system_admin":
            return Response(
                {"error": "Access denied. System admins only."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Get all AI statuses
        ai_statuses = AIStatus.objects.all()
        serializer = self.serializer_class(ai_statuses, many=True)
        
        # Return as a dict for easier frontend consumption
        status_dict = {}
        for status in serializer.data:
            feature_key = status['feature_name']
            status_dict[feature_key] = {
                'is_enabled': status['is_enabled'],
                'last_active': status['last_active']
            }
        
        return Response(status_dict)

    def create(self, request):
        """Toggle AI feature status"""
        # Check if user is system admin
        if request.user.role != "system_admin":
            return Response(
                {"error": "Access denied. System admins only."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        enabled = request.data.get("enabled")
        feature_name = request.data.get("feature_name", "admin_ai")  # Default to admin_ai for backward compatibility
        
        if enabled is None:
            return Response(
                {"error": "enabled field is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate feature_name
        valid_features = [choice[0] for choice in AIStatus.AI_FEATURE_CHOICES]
        if feature_name not in valid_features:
            return Response(
                {"error": f"Invalid feature_name. Must be one of: {valid_features}"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Toggle the specified AI feature
            ai_status = AIStatus.toggle_feature(
                feature_name=feature_name,
                enabled=enabled,
                user=request.user
            )
            
            serializer = self.serializer_class(ai_status)
            feature_display = ai_status.get_feature_name_display()
            
            return Response({
                "message": f"{feature_display} {'enabled' if enabled else 'disabled'} successfully",
                "status": serializer.data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Failed to toggle Admin AI: {str(e)}")
            return Response(
                {"error": "Failed to toggle AI status. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )





# Receptionist AI Chat View
@extend_schema(tags=["Public - Receptionist Chat"])
class ReceptionistChatView(viewsets.ViewSet):
    """
    Public receptionist AI chat functionality
    No authentication required - focused on platform information and guidance
    """
    permission_classes = []  # No authentication required
    serializer_class = ReceptionistChatMessageSerializer

    def get_queryset(self):
        """Get messages for a session (default session for simplicity)"""
        session_id = self.request.data.get('session_id', 'default')
        return ReceptionistChatMessage.objects.filter(session_id=session_id)

    def create(self, request):
        """Send a message and get AI response"""
        user_message = request.data.get("message", "").strip()
        session_id = request.data.get("session_id", "default")
        
        if not user_message:
            return Response(
                {"error": "Message cannot be empty"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Save user message
            user_chat_message = ReceptionistChatMessage.objects.create(
                session_id=session_id,
                sender="user",
                message=user_message
            )

            # Get last 5 messages for context (excluding the one we just saved)
            conversation_history = []
            recent_messages = self.get_queryset()[:5]
            
            for msg in recent_messages:
                if msg.id != user_chat_message.id:  # Skip the message we just saved
                    conversation_history.append({
                        "role": msg.api_role,
                        "content": msg.message
                    })

            # Get AI response using existing GroqService
            groq_service = GroqService()
            
            # Enhanced system prompt for receptionist-specific responses
            system_prompt = """You are Sana, the AI receptionist for Obeeoma, an AI-powered mental health platform tailored for Africa's workforce.

Your role is to:
1. Welcome visitors and explain what Obeeoma offers
2. Guide users about our mental health services and features
3. Explain how the platform connects to our mobile app
4. Direct users to create accounts for full access
5. Answer questions about mental health workplace wellness in Africa
6. Politely redirect conversations back to Obeeoma's services when asked off-topic

Key information about Obeeoma:
- AI-powered mental health platform for African workplaces
- Offers assessments, resources, and AI-guided support
- Connected to a mobile app for on-the-go access
- Focuses on empowering Africa's workforce mental wellness
- Provides confidential, accessible mental health support

Be warm, professional, and helpful. Always guide conversations toward how Obeeoma can help with mental health in the workplace. If asked questions outside our scope, politely relate it back to our mission or suggest creating an account to explore our full range of services."""
            
            # Add system prompt to conversation history
            full_conversation = [{"role": "system", "content": system_prompt}] + conversation_history
            
            ai_reply = groq_service.get_response(
                user_message=user_message,
                conversation_history=full_conversation,
            )

            # Save AI response
            ai_message = ReceptionistChatMessage.objects.create(
                session_id=session_id,
                sender="ai",
                message=ai_reply
            )

            # Return both messages
            response_data = {
                "user_message": ReceptionistChatMessageSerializer(user_chat_message).data,
                "ai_response": ReceptionistChatMessageSerializer(ai_message).data,
            }

            return Response(response_data, status=status.HTTP_201_CREATED)

        except ValueError as e:
            logger.error(f"Groq API configuration error: {str(e)}")
            return Response(
                {"error": "AI service not configured. Please contact support."},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        except Exception as e:
            logger.error(f"Receptionist AI chat error: {str(e)}")
            return Response(
                {"error": "Failed to process message. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
